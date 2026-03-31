"""
Interface web para exportar leads do banco de dados para Excel.

Requisitos:
    pip install flask psycopg2-binary openpyxl

Uso:
    python exportar_leads.py
    Acesse: http://localhost:5000
"""

import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path

import psycopg2
import psycopg2.extras
from flask import Flask, request, send_file, render_template_string, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATABASE_URL = "postgresql://neondb_owner:npg_FjPvo2f6dxRy@ep-mute-bird-ac4p4q9g-pooler.sa-east-1.aws.neon.tech/neondb?sslmode=require"
FORMATOS_FILE = Path("formatos_exportacao.json")
LIMITE_AUTO   = 1_000_000

NATUREZAS = {
    "2062": "Sociedade Empresária Limitada",
    "2070": "Sociedade Empresária em Nome Coletivo",
    "2089": "Sociedade Empresária em Comandita Simples",
    "2135": "Empresário Individual",
    "2232": "Sociedade Simples Pura",
    "2240": "Sociedade Simples Limitada",
    "2259": "Sociedade Simples em Nome Coletivo",
    "2267": "Sociedade Simples em Comandita Simples",
    "2305": "EIRELI",
    "2313": "EIRELI de Natureza Simples",
    "2127": "Sociedade em Conta de Participação",
    "2348": "Empresa Simples de Inovação",
    "4014": "Empresa Individual Imobiliária",
}

CAMPOS_DISPONIVEIS = {
    "razao_social":               "Razão Social",
    "cnpj":                       "CNPJ (14 dígitos)",
    "cnpj_numerico":              "CNPJ (só números)",
    "email":                      "E-mail",
    "telefone_principal":         "Telefone Principal",
    "telefone_principal_num":     "Telefone Principal (só números)",
    "telefone_secundario":        "Telefone Secundário",
    "telefone_secundario_num":    "Telefone Secundário (só números)",
    "atividade_principal_cod":    "CNAE Código",
    "atividade_principal_cod_num":"CNAE Código (só números)",
    "atividade_principal":        "CNAE Descrição",
    "natureza_juridica_cod":      "Natureza Jurídica Código",
    "natureza_juridica":          "Natureza Jurídica",
    "data_abertura":              "Data de Abertura",
    "ano_abertura":               "Ano de Abertura",
    "estado":                     "Estado (UF)",
    "cidade":                     "Cidade",
    "logradouro":                 "Logradouro",
    "bairro":                     "Bairro",
    "cep":                        "CEP",
    "nome_socio":                 "Nome do Sócio",
    "cpf_socio":                  "CPF do Sócio",
    "fixo_olos":                  "Fixo: OLOS",
    "fixo_flex":                  "Fixo: FLEX",
    "fixo_c6":                    "Fixo: C6",
    "fixo_vazio":                 "Vazio (null)",
}

FORMATOS_PADRAO = {
    "bernardo": {
        "nome": "Bernardo",
        "colunas": [
            {"header": "nome",   "campo": "razao_social"},
            {"header": "cpf",    "campo": "cnpj_numerico"},
            {"header": "livre1", "campo": "data_abertura"},
            {"header": "chave",  "campo": "email"},
            {"header": "livre3", "campo": "atividade_principal_cod_num"},
            {"header": "livre5", "campo": "fixo_vazio"},
            {"header": "livre7", "campo": "fixo_c6"},
            {"header": "fone1",  "campo": "telefone_principal_num"},
            {"header": "fone2",  "campo": "telefone_secundario_num"},
        ]
    },
    "olos": {
        "nome": "OLOS",
        "colunas": [
            {"header": "nome",   "campo": "razao_social"},
            {"header": "CNPJ",   "campo": "cnpj_numerico"},
            {"header": "livre1", "campo": "ano_abertura"},
            {"header": "EMAIL",  "campo": "email"},
            {"header": "livre3", "campo": "atividade_principal_cod_num"},
            {"header": "livre5", "campo": "fixo_olos"},
            {"header": "livre7", "campo": "fixo_flex"},
            {"header": "fone1",  "campo": "telefone_principal_num"},
            {"header": "fone2",  "campo": "telefone_secundario_num"},
        ]
    },
    "empresaaqui": {
        "nome": "EmpresaAqui",
        "colunas": [
            {"header": "razao",            "campo": "razao_social"},
            {"header": "cnpj",             "campo": "cnpj_numerico"},
            {"header": "data inicio ativ.","campo": "data_abertura"},
            {"header": "e-mail",           "campo": "email"},
            {"header": "cnae principal",   "campo": "atividade_principal_cod_num"},
            {"header": "livre5",           "campo": "fixo_vazio"},
            {"header": "livre7",           "campo": "fixo_c6"},
            {"header": "telefone 1",       "campo": "telefone_principal_num"},
            {"header": "telefone 2",       "campo": "telefone_secundario_num"},
        ]
    },
}

UFS = ["AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS","MT","PA",
       "PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC","SE","SP","TO"]

HTML = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Exportar Leads</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500&display=swap" rel="stylesheet">
<style>
:root{--bg:#f0ede8;--surface:#fff;--surface2:#f8f6f2;--border:#e0dcd4;--border2:#c8c3b8;--accent:#5b4fcf;--accent-l:#edeaff;--accent-d:#4338a8;--green:#0f7a56;--green-l:#e6f7f0;--text:#1c1917;--muted:#78716c;--muted2:#a8a29e;--danger:#be123c;--danger-l:#fff1f2;--warn:#92400e;--warn-l:#fffbeb;--radius:8px;--radius-lg:12px}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:'Inter',sans-serif;font-size:14px;line-height:1.5}
.layout{display:grid;grid-template-columns:240px 1fr;min-height:100vh}
.sidebar{background:var(--surface);border-right:1px solid var(--border);padding:0;position:sticky;top:0;height:100vh;overflow-y:auto;display:flex;flex-direction:column}
.sidebar-brand{padding:20px;border-bottom:1px solid var(--border)}
.sidebar-brand h1{font-family:'JetBrains Mono',monospace;font-size:13px;font-weight:500;color:var(--accent)}
.sidebar-brand p{font-size:11px;color:var(--muted2);margin-top:3px}
.nav-group{padding:16px 0 4px}
.nav-label{padding:0 16px 6px;font-size:10px;font-weight:600;color:var(--muted2);text-transform:uppercase;letter-spacing:1.2px;display:block}
.nav-btn{display:flex;align-items:center;gap:8px;padding:8px 16px;font-size:13px;color:var(--muted);cursor:pointer;border:none;background:none;width:100%;text-align:left;transition:all .15s;position:relative}
.nav-btn:hover{color:var(--text);background:var(--surface2)}
.nav-btn.active{color:var(--accent);background:var(--accent-l);font-weight:500}
.nav-btn.active::before{content:'';position:absolute;left:0;top:0;bottom:0;width:3px;background:var(--accent);border-radius:0 2px 2px 0}
.nav-icon{width:16px;height:16px;border-radius:4px;background:currentColor;opacity:.5;flex-shrink:0}
.main{padding:36px 40px;max-width:1100px;width:100%;box-sizing:border-box}
.page{display:none !important}.page.active{display:block !important}
.page-title{font-size:22px;font-weight:600;margin-bottom:3px}
.page-sub{color:var(--muted);font-size:13px;margin-bottom:28px}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:20px;margin-bottom:14px}
.section-title{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.9px;color:var(--muted);margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid var(--border)}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.g3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px}
.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}
.field-label{display:block;font-size:11px;font-weight:600;color:var(--muted);margin-bottom:5px;text-transform:uppercase;letter-spacing:.4px}
input[type=text],input[type=number],select:not([multiple]){width:100%;padding:8px 11px;border:1px solid var(--border);border-radius:6px;font-size:13px;font-family:inherit;background:var(--surface2);color:var(--text);outline:none;transition:border-color .15s}
input:focus,select:focus{border-color:var(--accent)}
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 18px;border-radius:6px;font-size:13px;font-weight:500;border:1px solid transparent;cursor:pointer;font-family:inherit;transition:all .15s}
.btn-primary{background:var(--accent);color:#fff}.btn-primary:hover{background:var(--accent-d)}
.btn-secondary{background:var(--surface2);color:var(--text);border-color:var(--border)}.btn-secondary:hover{border-color:var(--border2)}
.btn-lg{padding:11px 28px;font-size:14px}
.checkbox-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:6px}
.cb-item{display:flex;align-items:center;gap:8px;padding:7px 10px;border:1px solid var(--border);border-radius:6px;cursor:pointer;transition:border-color .15s;font-size:13px}
.cb-item:hover{border-color:var(--accent)}.cb-item input{accent-color:var(--accent);cursor:pointer;width:14px;height:14px}
.cb-code{font-family:'JetBrains Mono',monospace;font-size:10px;color:var(--green);margin-left:auto}
.quick-tags{display:flex;flex-wrap:wrap;gap:6px}
.qtag{padding:3px 10px;border:1px solid var(--border2);border-radius:20px;font-size:11px;color:var(--muted);cursor:pointer;transition:all .15s;background:var(--surface)}
.qtag:hover{border-color:var(--accent);color:var(--accent);background:var(--accent-l)}
.toggle-group{display:flex;gap:0}
.toggle-opt{padding:7px 14px;border:1px solid var(--border);font-size:12px;cursor:pointer;transition:all .15s;background:var(--surface2);color:var(--muted)}
.toggle-opt:first-child{border-radius:6px 0 0 6px}.toggle-opt:last-child{border-radius:0 6px 6px 0;border-left:0}
.toggle-opt.on{background:var(--accent-l);color:var(--accent);border-color:var(--accent);font-weight:500}
.toggle-opt input{display:none}
.split-box{background:var(--warn-l);border:1px solid #fde68a;border-radius:var(--radius);padding:14px;margin-top:8px}
.split-toggle{position:relative;width:36px;height:20px;cursor:pointer}
.split-toggle input{display:none}
.split-track{position:absolute;inset:0;background:var(--border2);border-radius:10px;transition:.2s}
.split-thumb{position:absolute;top:2px;left:2px;width:16px;height:16px;background:#fff;border-radius:50%;transition:.2s;box-shadow:0 1px 3px rgba(0,0,0,.2)}
.split-toggle input:checked~.split-track{background:var(--accent)}
.split-toggle input:checked~.split-thumb{left:18px}
.fmt-radio{display:flex;align-items:center;gap:12px;padding:12px 14px;border:2px solid var(--border);border-radius:8px;cursor:pointer;transition:all .15s}
.fmt-radio:hover{border-color:var(--border2)}.fmt-radio.sel{border-color:var(--accent);background:var(--accent-l)}
.fmt-radio input{accent-color:var(--accent);width:15px;height:15px}
.fmt-radio-name{font-weight:600;font-size:14px}.fmt-radio-desc{font-size:11px;color:var(--muted)}
.col-row{display:grid;grid-template-columns:28px 1fr 1.2fr 1fr 32px;gap:8px;align-items:center;background:var(--surface2);border:1px solid var(--border);border-radius:6px;padding:6px 8px;margin-bottom:6px}
.col-num{font-family:'JetBrains Mono',monospace;font-size:10px;color:var(--muted2);text-align:center}
.col-del{background:none;border:none;color:var(--muted2);cursor:pointer;font-size:18px;line-height:1;padding:0 4px;transition:color .15s}
.col-del:hover{color:var(--danger)}
/* Dashboard */
.dash-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:16px 20px;display:flex;flex-direction:column;gap:4px}
.dash-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:16px;margin-bottom:4px}
.dash-num{font-size:26px;font-weight:600;letter-spacing:-1px;color:var(--text)}
.dash-lbl{font-size:12px;color:var(--muted)}.dash-pct{font-size:11px;font-weight:500}
.bar-row{display:flex;flex-direction:column;gap:5px;margin-bottom:8px}
.bar-label{display:flex;justify-content:space-between;font-size:12px;color:var(--muted)}
.bar-track{height:8px;background:var(--surface2);border-radius:4px;overflow:hidden;border:1px solid var(--border)}
.bar-fill{height:100%;border-radius:4px;transition:width .8s cubic-bezier(.4,0,.2,1)}
/* API overlay */
.ainp{background:rgba(59,130,246,.04);border:1px solid rgba(59,130,246,.18);color:#4f8ef7;padding:7px 10px;border-radius:3px;font-family:'JetBrains Mono',monospace;font-size:11px;outline:none;box-sizing:border-box}
.ainp:focus{border-color:#4f8ef7}.ainp option{background:#080c1c;color:#4f8ef7}
.albl{color:rgba(59,130,246,.4);font-size:9px;margin-bottom:4px;letter-spacing:1px}
.atag{padding:3px 10px;border:1px solid rgba(59,130,246,.22);border-radius:20px;font-size:9px;color:rgba(59,130,246,.55);cursor:pointer;transition:all .15s}
.atag:hover{border-color:#4f8ef7;color:#4f8ef7}
.preset-btn{width:100%;padding:9px 12px;background:rgba(59,130,246,.04);border:1px solid rgba(59,130,246,.15);color:rgba(59,130,246,.7);font-family:'JetBrains Mono',monospace;font-size:10.5px;cursor:pointer;border-radius:3px;text-align:left;transition:all .15s}
.preset-btn:hover{background:rgba(59,130,246,.1);border-color:rgba(59,130,246,.5);color:#4f8ef7}
.preset-btn.active-preset{background:rgba(59,130,246,.15);border-color:#4f8ef7;color:#4f8ef7;box-shadow:inset 3px 0 0 #4f8ef7}
.aopt-item{display:flex;align-items:flex-start;gap:9px;cursor:pointer;padding:8px 10px;border:1px solid rgba(59,130,246,.1);border-radius:3px;transition:border-color .15s}
.aopt-item:hover{border-color:rgba(59,130,246,.3)}
.aopt-title{font-size:10.5px;color:rgba(59,130,246,.8);font-weight:500}
.aopt-desc{font-size:9px;color:rgba(59,130,246,.35);margin-top:2px;line-height:1.4}
.astat{background:rgba(0,0,0,.3);border:1px solid;border-radius:3px;padding:10px 14px}
.astat-lbl{font-size:9px;margin-bottom:3px;letter-spacing:1px}
.astat-val{font-size:24px;font-weight:500}
@keyframes pulse-dot{0%,100%{opacity:1;box-shadow:0 0 0 0 rgba(15,122,86,.4)}70%{box-shadow:0 0 0 6px rgba(15,122,86,0)}}
.preset-card{display:flex;flex-direction:column;align-items:center;gap:4px;padding:14px 8px;background:var(--surface);border:2px solid var(--border);border-radius:var(--radius-lg);cursor:pointer;transition:all .15s;font-family:inherit}
.preset-card:hover{border-color:var(--accent);background:var(--accent-l);transform:translateY(-2px);box-shadow:0 4px 12px rgba(91,79,207,.12)}
.preset-card.active-preset{border-color:var(--accent);background:var(--accent-l);box-shadow:0 0 0 3px rgba(91,79,207,.15)}
.preset-card-title{font-size:12px;font-weight:600;color:var(--text)}
.preset-card-desc{font-size:10px;color:var(--muted);text-align:center}
.aopt-row{display:flex;align-items:flex-start;gap:10px;cursor:pointer;padding:10px 12px;border:1px solid var(--border);border-radius:8px;transition:border-color .15s;background:var(--surface)}
.aopt-row:hover{border-color:var(--accent)}
@keyframes apidot{0%,100%{opacity:1;box-shadow:0 0 7px #4f8ef7}50%{opacity:.3;box-shadow:none}}
@keyframes api-enter{from{opacity:0;transform:scale(1.008) translateY(-8px)}to{opacity:1;transform:scale(1) translateY(0)}}
#api-overlay.show{animation:api-enter .35s cubic-bezier(.16,1,.3,1) forwards}
</style>
</head>
<body>
<div class="layout">
<nav class="sidebar">
  <div class="sidebar-brand"><h1>CNPJ → LEADS</h1><p>Base Receita Federal</p></div>
  <div class="nav-group">
    <span class="nav-label">Exportação</span>
    <button class="nav-btn active" onclick="goPage('filtros',this)"><span class="nav-icon" style="background:#5b4fcf"></span>Filtros &amp; Exportar</button>
  </div>
  <div class="nav-group">
    <span class="nav-label">Configuração</span>
    <button class="nav-btn" onclick="goPage('formatos',this)"><span class="nav-icon" style="background:#0f7a56"></span>Formatos</button>
    <button class="nav-btn" onclick="goPage('novo-formato',this)"><span class="nav-icon" style="background:#b45309"></span>Novo Formato</button>
  </div>
  <div class="nav-group">
    <span class="nav-label">API C6</span>
    <button class="nav-btn" onclick="entrarModoApi(this)"><span class="nav-icon" style="background:#b45309"></span>Consulta API</button>
    <button class="nav-btn" onclick="entrarModoApiResultados(this)"><span class="nav-icon" style="background:#1e6fa8"></span>Resultados</button>
    <button class="nav-btn" onclick="goPage('teste-fish',this)"><span class="nav-icon" style="background:#0f7a56"></span>Teste Fish</button>
  </div>
</nav>
<div class="main">

<!-- FILTROS -->
<div id="page-filtros" class="page active">

  <!-- HEADER -->
  <div style="margin-bottom:28px">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">
      <div style="width:36px;height:36px;border-radius:10px;background:linear-gradient(135deg,var(--accent),#7c6fe0);display:flex;align-items:center;justify-content:center;font-size:18px">📤</div>
      <div>
        <div class="page-title" style="margin:0">Filtros &amp; Exportar</div>
        <div class="page-sub" style="margin:0">Configure os critérios e exporte para Excel</div>
      </div>
    </div>
  </div>

  <form id="form-export" onsubmit="return false;">

    <!-- PRESETS -->
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;padding:10px 14px;background:var(--surface2);border:1px solid var(--border);border-radius:8px">
      <span style="font-size:11px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;white-space:nowrap">Presets:</span>
      <div id="preset-list" style="display:flex;gap:6px;flex-wrap:wrap;flex:1"></div>
      <button type="button" onclick="salvarPreset()" class="btn btn-secondary" style="font-size:11px;padding:5px 12px;white-space:nowrap">💾 Salvar atual</button>
    </div>

    <!-- PERÍODO + LIMPEZA lado a lado -->
    <div class="g2" style="margin-bottom:14px;align-items:start">
      <!-- PERÍODO E LOCALIZAÇÃO -->
      <div class="card">
            <div class="section-title">📍 Período e Localização</div>
            <div style="display:grid;grid-template-columns:1fr 1fr 2fr;gap:16px">
              <div>
                <label class="field-label">ANO DE — DE</label>
                <input type="number" name="ano_de" min="1900" max="2026" placeholder="ex: 2020" value="{{ f.ano_de }}" style="width:100%">
              </div>
              <div>
                <label class="field-label">ANO DE — ATÉ</label>
                <input type="number" name="ano_ate" min="1900" max="2026" placeholder="ex: 2025" value="{{ f.ano_ate }}" style="width:100%">
              </div>
              <div>
                <label class="field-label">ESTADO (UF)</label>
                <label style="display:flex;align-items:center;gap:8px;padding:8px 12px;background:var(--surface2);border:1px solid var(--border);border-radius:6px;cursor:pointer;font-size:13px;font-weight:500">
                  <input type="checkbox" id="uf-tudo" checked style="accent-color:var(--accent);width:15px;height:15px" onchange="toggleUFPanel(this)">
                  🌎 Todos os estados (Brasil inteiro)
                </label>
                <div id="uf-panel" style="display:none;margin-top:8px">
                  <div class="quick-tags" style="margin-bottom:8px">
                    <span class="qtag" onclick="toggleUFs(true)">Marcar todos</span>
                    <span class="qtag" onclick="toggleUFs(false)">Limpar</span>
                    <span class="qtag" onclick="toggleUFGroup(['SP','RJ','MG','ES'])">Sudeste</span>
                    <span class="qtag" onclick="toggleUFGroup(['RS','SC','PR'])">Sul</span>
                    <span class="qtag" onclick="toggleUFGroup(['BA','SE','AL','PE','PB','RN','CE','PI','MA'])">Nordeste</span>
                    <span class="qtag" onclick="toggleUFGroup(['GO','MT','MS','DF'])">C-Oeste</span>
                    <span class="qtag" onclick="toggleUFGroup(['AM','PA','AC','RO','RR','AP','TO'])">Norte</span>
                  </div>
                  <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(58px,1fr));gap:4px">
                    {% for uf in ufs %}
                    <label style="display:flex;align-items:center;gap:4px;padding:4px 8px;background:var(--surface2);border:1px solid var(--border);border-radius:5px;cursor:pointer;font-size:12px;font-weight:500;transition:all .15s" class="uf-item">
                      <input type="checkbox" name="uf" value="{{ uf }}" checked style="accent-color:var(--accent);width:12px;height:12px" onchange="updateUFStyle(this)">{{ uf }}
                    </label>
                    {% endfor %}
                  </div>
                </div>
              </div>
            </div>
      </div>

      <!-- LIMPEZA E ENRIQUECIMENTO -->
      <div class="card">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px">
              <div class="section-title" style="margin:0">🧹 Limpeza &amp; Enriquecimento</div>
              <label style="display:flex;align-items:center;gap:8px;cursor:pointer;padding:5px 12px;border:1px solid var(--accent);border-radius:20px;background:var(--accent-l)">
                <input type="checkbox" name="limpeza_ativo" id="limpeza_ativo" value="sim" style="accent-color:var(--accent);width:13px;height:13px" onchange="toggleLimpeza(this)">
                <span style="font-size:12px;font-weight:600;color:var(--accent)">Ativar pipeline</span>
              </label>
            </div>
      
            <div id="limpeza-opts" style="display:none">
              <!-- Preset rápido -->
              <div style="display:flex;gap:8px;margin-bottom:14px">
                <button type="button" onclick="limpezaPreset('completo')" class="btn btn-secondary" style="font-size:11px;padding:5px 12px">⚡ Processo completo</button>
                <button type="button" onclick="limpezaPreset('blocklist')" class="btn btn-secondary" style="font-size:11px;padding:5px 12px">🚫 Só blocklist</button>
                <button type="button" onclick="limpezaPreset('enrich')" class="btn btn-secondary" style="font-size:11px;padding:5px 12px">📞 Só enriquecer</button>
                <button type="button" onclick="limpezaPreset('nenhum')" class="btn btn-secondary" style="font-size:11px;padding:5px 12px">✕ Limpar</button>
              </div>
      
              <!-- 4 opções lado a lado -->
              <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px">
      
                <label style="display:flex;flex-direction:column;gap:6px;padding:12px;border:2px solid var(--border);border-radius:8px;cursor:pointer;transition:border-color .15s" class="limpeza-opt" id="opt-raiz">
                  <div style="display:flex;align-items:center;gap:8px">
                    <input type="checkbox" name="limpeza_raiz" id="limpeza_raiz" value="sim" style="accent-color:var(--accent);width:14px;height:14px" onchange="updateLimpezaOpt(this,'opt-raiz')">
                    <span style="font-size:13px;font-weight:600">📋 Raiz</span>
                  </div>
                  <span style="font-size:11px;color:var(--muted);line-height:1.4">Remove CNPJs presentes na tabela <code>raiz_cnpjs</code> (evita duplicar clientes já trabalhados)</span>
                  <span style="font-size:10px;color:var(--accent);font-weight:500">Etapa 1</span>
                </label>
      
                <label style="display:flex;flex-direction:column;gap:6px;padding:12px;border:2px solid var(--border);border-radius:8px;cursor:pointer;transition:border-color .15s" class="limpeza-opt" id="opt-enrich">
                  <div style="display:flex;align-items:center;gap:8px">
                    <input type="checkbox" name="limpeza_enrich" id="limpeza_enrich" value="sim" style="accent-color:var(--green);width:14px;height:14px" onchange="updateLimpezaOpt(this,'opt-enrich')">
                    <span style="font-size:13px;font-weight:600">📞 Enriquecer</span>
                  </div>
                  <span style="font-size:11px;color:var(--muted);line-height:1.4">Busca telefones adicionais no banco externo e adiciona nas colunas livres (append)</span>
                  <span style="font-size:10px;color:var(--green);font-weight:500">Etapa 2</span>
                </label>
      
                <label style="display:flex;flex-direction:column;gap:6px;padding:12px;border:2px solid var(--border);border-radius:8px;cursor:pointer;transition:border-color .15s" class="limpeza-opt" id="opt-fixos">
                  <div style="display:flex;align-items:center;gap:8px">
                    <input type="checkbox" name="limpeza_fixos" id="limpeza_fixos" value="sim" style="accent-color:var(--warn);width:14px;height:14px" onchange="updateLimpezaOpt(this,'opt-fixos')">
                    <span style="font-size:13px;font-weight:600">📵 Fixos</span>
                  </div>
                  <span style="font-size:11px;color:var(--muted);line-height:1.4">Remove telefones fixos (3º dígito ≤ 5). Linha permanece, célula fica vazia</span>
                  <span style="font-size:10px;color:var(--warn);font-weight:500">Etapa 3</span>
                </label>
      
                <label style="display:flex;flex-direction:column;gap:6px;padding:12px;border:2px solid var(--border);border-radius:8px;cursor:pointer;transition:border-color .15s" class="limpeza-opt" id="opt-blocklist">
                  <div style="display:flex;align-items:center;gap:8px">
                    <input type="checkbox" name="limpeza_blocklist" id="limpeza_blocklist" value="sim" style="accent-color:var(--danger);width:14px;height:14px" onchange="updateLimpezaOpt(this,'opt-blocklist')">
                    <span style="font-size:13px;font-weight:600">🚫 Blocklist</span>
                  </div>
                  <span style="font-size:11px;color:var(--muted);line-height:1.4">Remove fones que estão na tabela <code>blocklist</code> do banco externo (em lotes de 30k)</span>
                  <span style="font-size:10px;color:var(--danger);font-weight:500">Etapa 4</span>
                </label>
      
              </div>
            </div>
      
            <div id="limpeza-info" style="font-size:12px;color:var(--muted);margin-top:10px">
              Ative o pipeline para configurar as etapas de limpeza da planilha exportada.
            </div>
          </div>
    </div>

    <!-- ATIVIDADE E CONTATO -->
    <div class="card" style="margin-bottom:14px">
      <div class="section-title">🎯 Atividade, Contato e Filtros</div>
      <div class="g2" style="margin-bottom:14px">
        <div>
          <label class="field-label">CNAE PRINCIPAL</label>
          <input type="text" name="cnaes" placeholder="ex: 6201500, 4781400" value="{{ f.cnaes }}">
        </div>
        <div>
          <label class="field-label">LIMITE DE REGISTROS</label>
          <input type="number" name="limite" min="1" max="50000000" placeholder="sem limite (tudo)" value="{{ f.limite }}">
        </div>
      </div>
      <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px">
        <div>
          <label class="field-label">TELEFONE</label>
          <div class="toggle-group">
            <label class="toggle-opt {% if f.com_telefone=='nao' %}on{% endif %}"><input type="radio" name="com_telefone" value="nao" {% if f.com_telefone=='nao' %}checked{% endif %} onchange="updateToggle(this)">Todos</label>
            <label class="toggle-opt {% if f.com_telefone=='sim' %}on{% endif %}"><input type="radio" name="com_telefone" value="sim" {% if f.com_telefone=='sim' %}checked{% endif %} onchange="updateToggle(this)">Só c/ tel.</label>
          </div>
        </div>
        <div>
          <label class="field-label">E-MAIL</label>
          <div class="toggle-group">
            <label class="toggle-opt {% if f.com_email=='nao' %}on{% endif %}"><input type="radio" name="com_email" value="nao" {% if f.com_email=='nao' %}checked{% endif %} onchange="updateToggle(this)">Todos</label>
            <label class="toggle-opt {% if f.com_email=='sim' %}on{% endif %}"><input type="radio" name="com_email" value="sim" {% if f.com_email=='sim' %}checked{% endif %} onchange="updateToggle(this)">Só c/ email</label>
          </div>
        </div>
        <div>
          <label class="field-label">MEI</label>
          <div class="toggle-group">
            <label class="toggle-opt {% if f.sem_mei=='nao' %}on{% endif %}"><input type="radio" name="sem_mei" value="nao" {% if f.sem_mei=='nao' %}checked{% endif %} onchange="updateToggle(this)">Incluir</label>
            <label class="toggle-opt {% if f.sem_mei=='sim' %}on{% endif %}"><input type="radio" name="sem_mei" value="sim" {% if f.sem_mei=='sim' %}checked{% endif %} onchange="updateToggle(this)">Excluir MEI</label>
          </div>
        </div>
        <div>
          <label class="field-label">STATUS API C6</label>
          <div class="toggle-group">
            <label class="toggle-opt {% if f.so_disponiveis=='nao' %}on{% endif %}"><input type="radio" name="so_disponiveis" value="nao" {% if f.so_disponiveis=='nao' %}checked{% endif %} onchange="updateToggle(this)">Todos</label>
            <label class="toggle-opt {% if f.so_disponiveis=='sim' %}on{% endif %}"><input type="radio" name="so_disponiveis" value="sim" {% if f.so_disponiveis=='sim' %}checked{% endif %} onchange="updateToggle(this)">Disponíveis</label>
          </div>
        </div>
      </div>
    </div>

    <!-- NATUREZA JURÍDICA -->
    <div class="card" style="margin-bottom:14px">
      <div class="section-title">⚖ Natureza Jurídica</div>
      <div class="quick-tags" style="margin-bottom:10px">
        <span class="qtag" onclick="toggleAll(true)">Selecionar todas</span>
        <span class="qtag" onclick="toggleAll(false)">Limpar</span>
        <span class="qtag" onclick="toggleGroup(['2135','2305','2313','2348','4014'])">Individuais</span>
        <span class="qtag" onclick="toggleGroup(['2062','2232','2240','2259','2267','2070','2089'])">Sociedades</span>
      </div>
      <div class="checkbox-grid">
        {% for cod, nome in naturezas.items() %}
        <label class="cb-item">
          <input type="checkbox" name="naturezas" value="{{ cod }}" {% if cod in f.naturezas_sel %}checked{% endif %}>
          <span>{{ nome }}</span><span class="cb-code">{{ cod }}</span>
        </label>
        {% endfor %}
      </div>
    </div>

    <!-- FORMATO E DIVISÃO em 2 colunas -->
    <div class="g2" style="margin-bottom:14px">
      <div class="card">
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px">
          <div class="section-title" style="margin:0">📋 Formato de Exportação</div>
          <label style="display:flex;align-items:center;gap:7px;cursor:pointer;font-size:12px;color:var(--muted)">
            <input type="checkbox" id="dual-fmt-toggle" style="accent-color:var(--accent);width:13px;height:13px" onchange="toggleDualFmt(this)">
            Dividir em 2 formatos
          </label>
        </div>

        <!-- FORMATO ÚNICO -->
        <div id="fmt-single">
          <div style="display:flex;flex-direction:column;gap:8px">
            {% for key, fmt in formatos.items() %}
            <label class="fmt-radio {% if f.formato==key %}sel{% endif %}" onclick="selectFmt(this)">
              <input type="radio" name="formato" value="{{ key }}" {% if f.formato==key %}checked{% endif %}>
              <div><div class="fmt-radio-name">{{ fmt.nome }}</div>
              <div class="fmt-radio-desc">{{ fmt.colunas|length }} col · {% for c in fmt.colunas[:3] %}{{ c.header }}{% if not loop.last %}, {% endif %}{% endfor %}...</div></div>
            </label>
            {% endfor %}
          </div>
        </div>

        <!-- DUPLO FORMATO -->
        <div id="fmt-dual" style="display:none;flex-direction:column;gap:12px">
          <input type="hidden" name="dual_fmt_ativo" id="dual_fmt_ativo" value="nao">

          <!-- Modo de divisão -->
          <div style="display:flex;align-items:center;gap:10px;padding:10px 12px;background:var(--accent-l);border:1px solid var(--accent);border-radius:8px">
            <span style="font-size:12px;font-weight:500;color:var(--accent)">Modo de divisão:</span>
            <div class="toggle-group">
              <label class="toggle-opt on" id="modo-qtd-lbl"><input type="radio" name="modo_divisao" value="qtd" checked onchange="updateToggle(this);toggleModoDivisao('qtd')">Quantidade fixa</label>
              <label class="toggle-opt" id="modo-pct-lbl"><input type="radio" name="modo_divisao" value="pct" onchange="updateToggle(this);toggleModoDivisao('pct')">Porcentagem (%)</label>
            </div>
            <span id="info-pct" style="font-size:11px;color:var(--muted);display:none">O sistema conta o total e calcula automaticamente</span>
          </div>

          {% for slot in ['a','b'] %}
          <div style="border:1px solid var(--border);border-radius:8px;padding:12px;background:var(--surface2)">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:10px">
              <span style="font-size:12px;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.5px">
                {% if slot == 'a' %}📗 Lista A{% else %}📘 Lista B{% endif %}
              </span>
            </div>
            <div class="g2" style="margin-bottom:8px">
              <div>
                <label class="field-label">NOME DO ARQUIVO</label>
                <input type="text" name="nome_dual_{{ slot }}" placeholder="ex: {% if slot == 'a' %}recentes{% else %}antigos{% endif %}" style="width:100%">
              </div>
              <div>
                <!-- modo quantidade -->
                <div class="div-modo-qtd">
                  <label class="field-label">QTD DE REGISTROS</label>
                  <input type="number" name="qtd_dual_{{ slot }}" placeholder="ex: 500000" min="1" style="width:100%">
                </div>
                <!-- modo porcentagem -->
                <div class="div-modo-pct" style="display:none">
                  <label class="field-label">PORCENTAGEM (%)</label>
                  <div style="display:flex;align-items:center;gap:6px">
                    <input type="number" name="pct_dual_{{ slot }}" placeholder="{% if slot == 'a' %}50{% else %}50{% endif %}" min="1" max="100" step="1" style="width:80px" oninput="atualizarSomaPct()">
                    <span style="font-size:13px;font-weight:500">%</span>
                  </div>
                </div>
              </div>
            </div>
            <div class="g2">
              <div>
                <label class="field-label">ORDENAÇÃO POR DATA</label>
                <div class="toggle-group">
                  <label class="toggle-opt on"><input type="radio" name="ordem_dual_{{ slot }}" value="recentes" checked onchange="updateToggle(this)">Mais recentes</label>
                  <label class="toggle-opt"><input type="radio" name="ordem_dual_{{ slot }}" value="antigos" onchange="updateToggle(this)">Mais antigos</label>
                </div>
              </div>
              <div>
                <label class="field-label">FORMATO</label>
                <select name="formato_dual_{{ slot }}" style="width:100%">
                  {% for key, fmt in formatos.items() %}
                  <option value="{{ key }}">{{ fmt.nome }} ({{ fmt.colunas|length }} col)</option>
                  {% endfor %}
                </select>
              </div>
            </div>
          </div>
          {% endfor %}

          <div id="aviso-pct" style="display:none;padding:8px 12px;background:var(--warn-l);border:1px solid #fde68a;border-radius:6px;font-size:12px;color:var(--warn)"></div>
        </div>
      </div>
      <div class="card">
        <div class="section-title">✂ Divisão em Listas</div>
        <div style="margin-bottom:14px">
          <label class="field-label">NOME DA LISTA</label>
          <input type="text" name="nome_lista" id="nome_lista" placeholder="ex: leads_sp_2024" value="{{ f.get('nome_lista','') }}" style="width:100%" oninput="atualizarPreviewNome()">
          <div id="preview-nome" style="margin-top:6px;font-size:11px;color:var(--muted);padding:5px 9px;background:var(--surface2);border:1px solid var(--border);border-radius:5px;display:none"></div>
        </div>
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px">
          <label class="split-toggle"><input type="checkbox" id="split-toggle" name="split_ativo" value="sim" onchange="toggleSplit();atualizarPreviewNome()"><div class="split-track"></div><div class="split-thumb"></div></label>
          <span style="font-size:13px;font-weight:500">Dividir em múltiplas listas</span>
        </div>
        <div id="split-options" style="display:none">
          <div style="background:var(--warn-l);border:1px solid #fde68a;border-radius:var(--radius);padding:12px;margin-bottom:10px">
            <p style="font-size:12px;color:var(--warn)">⚠ Resultado dividido em arquivos .xlsx dentro de um .zip.</p>
          </div>
          <label class="field-label">REGISTROS POR LISTA</label>
          <input type="number" name="split_qtd" min="1000" max="5000000" placeholder="ex: 500000" value="{{ f.split_qtd }}" style="width:100%">
        </div>
        <div id="split-auto-info">
          <p style="font-size:12px;color:var(--muted)">Sem divisão: gera um único arquivo.</p>
        </div>
        <div style="margin-top:14px;padding-top:14px;border-top:1px solid var(--border)">
          <label class="field-label">LINHAS POR LOTE NO EXCEL</label>
          <div style="display:flex;align-items:center;gap:8px">
            <input type="number" name="batch_excel" min="10000" max="1000000" placeholder="padrão: 200000" value="{{ f.batch_excel }}" style="width:160px">
            <span style="font-size:11px;color:var(--muted)">Maior = mais rápido, mais RAM</span>
          </div>
        </div>
      </div>
    </div>

    <!-- BOTÃO -->
    <button type="button" onclick="iniciarExportacao()" class="btn btn-primary btn-lg" id="btn-exportar" style="width:100%;padding:14px;font-size:14px;letter-spacing:.3px;box-shadow:0 4px 14px rgba(91,79,207,.25);transition:all .2s" onmouseover="this.style.transform='translateY(-1px)';this.style.boxShadow='0 6px 20px rgba(91,79,207,.35)'" onmouseout="this.style.transform='';this.style.boxShadow='0 4px 14px rgba(91,79,207,.25)'">
      ⬇ Exportar Excel
    </button>

    <!-- LOG -->
    <div id="log-panel" style="display:none;margin-top:20px" class="card">
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px">
        <div style="display:flex;align-items:center;gap:8px">
          <span style="width:8px;height:8px;border-radius:50%;background:var(--green);display:inline-block;animation:pulse-dot 2s infinite"></span>
          <div class="section-title" style="margin:0">Progresso</div>
        </div>
        <span id="log-status" style="font-size:11px;color:var(--muted)"></span>
        <div id="download-btn-area" style="margin-left:auto"></div>
      </div>
      <div id="log-box" style="background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);padding:12px;font-family:'JetBrains Mono',monospace;font-size:11px;max-height:360px;overflow-y:auto;white-space:pre-wrap;line-height:1.6;color:var(--text)"></div>
    </div>
  </form>
</div>


<!-- FORMATOS -->
<div id="page-formatos" class="page">
  <div class="page-title">Formatos de Exportação</div>
  <div class="page-sub">Gerencie os modelos de colunas disponíveis</div>
  {% for key, fmt in formatos.items() %}
  <div class="card">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px">
      <span style="font-size:15px;font-weight:600">{{ fmt.nome }}</span>
      <span style="padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500;background:var(--accent-l);color:var(--accent)">{{ fmt.colunas|length }} colunas</span>
      {% if key in ['bernardo','olos','empresaaqui'] %}<span style="padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500;background:var(--green-l);color:var(--green)">padrão</span>{% endif %}
      <div style="margin-left:auto;display:flex;gap:6px">
        <a href="/editar-formato/{{ key }}" class="btn btn-secondary" style="font-size:12px;padding:6px 12px">Editar</a>
        {% if key not in ['bernardo','olos','empresaaqui'] %}<a href="/excluir-formato/{{ key }}" class="btn" style="background:#fef2f2;color:var(--danger);border-color:#fecaca;font-size:12px;padding:6px 12px" onclick="return confirm('Excluir?')">Excluir</a>{% endif %}
      </div>
    </div>
    <div style="display:flex;flex-wrap:wrap;gap:5px">
      {% for col in fmt.colunas %}
      <div style="padding:3px 10px;background:var(--surface2);border:1px solid var(--border);border-radius:4px;font-size:11px"><span style="font-weight:600">{{ col.header }}</span><span style="color:var(--muted);margin-left:4px">← {{ col.campo }}</span></div>
      {% endfor %}
    </div>
  </div>
  {% endfor %}
</div>

<!-- NOVO FORMATO -->
<div id="page-novo-formato" class="page">
  <div class="page-title" id="titulo-novo">Criar Novo Formato</div>
  <div class="page-sub">Defina nome e colunas do novo modelo</div>
  <div class="card" style="padding:14px 16px;background:var(--warn-l);border-color:#fde68a;margin-bottom:16px">
    <p style="font-size:12px;color:var(--warn)">ℹ Cada coluna tem um <b>cabeçalho</b> (nome no Excel) e um <b>campo</b> (dado do banco). Use "Valor fixo manual" para texto fixo.</p>
  </div>
  <form method="POST" action="/salvar-formato" id="form-formato">
    <input type="hidden" name="formato_key" id="formato_key" value="">
    <div class="card">
      <div class="section-title">Identificação</div>
      <div class="g2">
        <div><label class="field-label">NOME DO FORMATO</label><input type="text" name="formato_nome" id="formato_nome" placeholder="ex: Meu CRM"></div>
        <div><label class="field-label">IDENTIFICADOR (sem espaços)</label><input type="text" name="formato_id" id="formato_id" placeholder="ex: meu_crm" pattern="[a-z0-9_]+"></div>
      </div>
    </div>
    <div class="card">
      <div class="section-title">Colunas</div>
      <div id="col-builder"></div>
      <div style="margin-top:10px"><button type="button" class="btn btn-secondary" onclick="addCol()">+ Adicionar Coluna</button></div>
    </div>
    <div style="display:flex;gap:8px">
      <button type="submit" class="btn btn-primary">Salvar Formato</button>
      <button type="button" class="btn btn-secondary" onclick="goPage('formatos',document.querySelector('.nav-btn:nth-child(4)'))">Cancelar</button>
    </div>
  </form>
</div>

<!-- API RESULTADOS DASHBOARD -->
<div id="page-api-resultados" class="page">
  <div class="page-title">Dashboard — Limpeza API</div>
  <div class="page-sub">Análise dos resultados da consulta C6 Bank</div>
  <div class="g4" style="margin-bottom:20px">
    <div class="dash-card"><div class="dash-icon" style="background:rgba(91,79,207,.1);color:var(--accent)">🏢</div><div class="dash-num" id="dc-total-bd" style="color:var(--muted2);font-size:18px">clique em buscar</div><div class="dash-lbl">Total no banco</div></div>
    <div class="dash-card"><div class="dash-icon" style="background:rgba(91,79,207,.1);color:var(--accent)">🔍</div><div class="dash-num" id="dc-processados">—</div><div class="dash-lbl">Processados</div><div class="dash-pct" id="dc-processados-pct" style="color:var(--accent)"></div></div>
    <div class="dash-card"><div class="dash-icon" style="background:rgba(15,122,86,.1);color:var(--green)">✅</div><div class="dash-num" id="dc-disp" style="color:var(--green)">—</div><div class="dash-lbl">Disponíveis</div><div class="dash-pct" id="dc-disp-pct" style="color:var(--green)"></div></div>
    <div class="dash-card"><div class="dash-icon" style="background:rgba(190,18,60,.08);color:var(--danger)">⛔</div><div class="dash-num" id="dc-cli" style="color:var(--danger)">—</div><div class="dash-lbl">Já clientes C6</div><div class="dash-pct" id="dc-cli-pct" style="color:var(--danger)"></div></div>
  </div>
  <div class="g2" style="margin-bottom:16px">
    <div class="card" style="padding:20px"><div class="section-title">Cobertura da base</div>
      <div style="display:flex;align-items:center;gap:24px">
        <canvas id="chart-cobertura" width="140" height="140" style="flex-shrink:0"></canvas>
        <div id="legend-cobertura" style="display:flex;flex-direction:column;gap:8px;font-size:13px"></div>
      </div>
    </div>
    <div class="card" style="padding:20px"><div class="section-title">Resultado da limpeza</div>
      <div style="display:flex;align-items:center;gap:24px">
        <canvas id="chart-resultado" width="140" height="140" style="flex-shrink:0"></canvas>
        <div id="legend-resultado" style="display:flex;flex-direction:column;gap:8px;font-size:13px"></div>
      </div>
    </div>
  </div>
  <div class="card" style="padding:20px;margin-bottom:16px">
    <div class="section-title">Taxa de conversão</div>
    <div id="dash-bars"></div>
  </div>

  <!-- NATUREZAS BREAKDOWN -->
  <div class="card" style="padding:20px;margin-bottom:16px">
    <div class="section-title">Natureza Jurídica — cobertura e conversão</div>
    <div id="dash-naturezas" style="overflow-x:auto">
      <p style="color:var(--muted);font-size:13px">Clique em 🔄 Atualizar para carregar.</p>
    </div>
  </div>
  <div class="card" style="padding:20px">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px">
      <div class="section-title" style="margin:0">Registros recentes</div>
      <div style="display:flex;gap:8px">
        <select id="res-filtro" style="padding:6px 10px;border:1px solid var(--border);border-radius:6px;font-size:12px;background:var(--surface2);color:var(--text)">
          <option value="">Todos</option><option value="disponivel">Disponível</option><option value="cliente">Cliente</option>
        </select>
        <input type="number" id="res-limite" value="50" min="1" max="500" style="width:70px;padding:6px 8px;border:1px solid var(--border);border-radius:6px;font-size:12px;background:var(--surface2);color:var(--text)">
        <button class="btn btn-secondary" onclick="carregarResultados()" style="font-size:12px;padding:6px 14px">🔄 Atualizar</button>
        <button class="btn btn-primary" onclick="exportarResultados()" style="font-size:12px;padding:6px 14px">⬇ Excel</button>
      </div>
    </div>
    <div id="res-tabela" style="overflow-x:auto;font-size:12px"></div>
  </div>
</div>

<!-- CONSULTA API - normal page -->
<div id="page-api" class="page">

  <!-- HEADER -->
  <div style="display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:28px">
    <div>
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
        <div style="width:36px;height:36px;border-radius:10px;background:linear-gradient(135deg,#b45309,#d97706);display:flex;align-items:center;justify-content:center;font-size:18px">🏦</div>
        <div>
          <div class="page-title" style="margin:0">Consulta API C6</div>
          <div class="page-sub" style="margin:0">Classifica CNPJs como disponível ou cliente C6 Bank</div>
        </div>
      </div>
    </div>
    <div style="display:flex;align-items:center;gap:8px;padding:6px 12px;background:var(--warn-l);border:1px solid #fde68a;border-radius:8px;font-size:12px;color:var(--warn)">
      ⛔ MEI sempre excluído
    </div>
  </div>

  <!-- PRESET CARDS -->
  <div style="margin-bottom:20px">
    <div style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.8px;color:var(--muted);margin-bottom:10px">Selecione um perfil de consulta</div>
    <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px">
      <button onclick="aplicarPreset('recentes')" id="preset-recentes" class="preset-card">
        <span style="font-size:22px">📅</span>
        <div class="preset-card-title">Recentes</div>
        <div class="preset-card-desc">Últimos 2 anos</div>
      </button>
      <button onclick="aplicarPreset('ativos')" id="preset-ativos" class="preset-card active-preset">
        <span style="font-size:22px">✅</span>
        <div class="preset-card-title">Toda base</div>
        <div class="preset-card-desc">Ativos sem MEI</div>
      </button>
      <button onclick="aplicarPreset('ei_recente')" id="preset-ei_recente" class="preset-card">
        <span style="font-size:22px">👤</span>
        <div class="preset-card-title">EI Recente</div>
        <div class="preset-card-desc">Empresário Ind.</div>
      </button>
      <button onclick="aplicarPreset('ltda')" id="preset-ltda" class="preset-card">
        <span style="font-size:22px">🏢</span>
        <div class="preset-card-title">Sociedades</div>
        <div class="preset-card-desc">Ltda e similares</div>
      </button>
      <button onclick="aplicarPreset('custom')" id="preset-custom" class="preset-card">
        <span style="font-size:22px">⚙</span>
        <div class="preset-card-title">Personalizado</div>
        <div class="preset-card-desc">Filtros manuais</div>
      </button>
    </div>
  </div>

  <!-- FILTROS CUSTOM -->
  <div id="aapi-filtros-avancados" style="display:none;margin-bottom:16px">
    <div class="card">
      <div class="section-title">Filtros Personalizados</div>
      <div class="g3" style="margin-bottom:14px">
        <div><label class="field-label">ANO DE</label><input id="aapi_ano_de" type="number" placeholder="2020"></div>
        <div><label class="field-label">ANO ATÉ</label><input id="aapi_ano_ate" type="number" placeholder="2024"></div>
        <div><label class="field-label">CNAE</label><input id="aapi_cnaes" type="text" placeholder="6201500, 4781400"></div>
      </div>
      <div>
        <label class="field-label">NATUREZA JURÍDICA</label>
        <div class="quick-tags" style="margin-bottom:8px">
          <span class="qtag" onclick="toggleAApiNat(true)">Todas</span>
          <span class="qtag" onclick="toggleAApiNat(false)">Limpar</span>
          <span class="qtag" onclick="toggleAApiNatGrp2(['2135','2305','2313'])">Individuais</span>
          <span class="qtag" onclick="toggleAApiNatGrp2(['2062','2232','2240','2259','2267','2070','2089'])">Sociedades</span>
        </div>
        <div class="checkbox-grid">
          {% for cod, nome in naturezas.items() %}
          <label class="cb-item">
            <input type="checkbox" class="aapi-nat" value="{{ cod }}" checked>
            <span>{{ nome }}</span><span class="cb-code">{{ cod }}</span>
          </label>
          {% endfor %}
        </div>
      </div>
    </div>
  </div>

  <!-- CONFIG + OPCOES em 2 colunas -->
  <div class="g2" style="margin-bottom:16px">
    <div class="card">
      <div class="section-title">Configuração da API</div>
      <div style="margin-bottom:12px">
        <label class="field-label">MODO DE CHAVE</label>
        <select id="aapi_key_mode" onchange="atualizarInfoLote(this)" style="width:100%">
          <option value="chave1">Chave 1 — Padrão</option>
          <option value="chave2">Chave 2 — Alternativa</option>
          <option value="intercalar">Intercalar — alterna 1 e 2</option>
          <option value="dupla">Dupla — simultâneo (recomendado)</option>
        </select>
        <div id="info-lote" style="margin-top:6px;font-size:11px;color:var(--muted);padding:6px 10px;background:var(--surface2);border-radius:6px;border:1px solid var(--border)">
          📦 Lote: 20.000 CNPJs por execução
        </div>
      </div>
      <div>
        <label class="field-label">DELAY ENTRE LOTES</label>
        <div style="display:flex;align-items:center;gap:8px">
          <input id="aapi_delay" type="number" value="3" min="0" step="0.5" style="width:80px">
          <span style="font-size:12px;color:var(--muted)">minutos</span>
        </div>
      </div>
    </div>

    <div class="card">
      <div class="section-title">Opções</div>
      <div style="display:flex;flex-direction:column;gap:8px">
        <label class="aopt-row">
          <input type="checkbox" id="aapi_reprocessar" style="accent-color:var(--accent);width:14px;height:14px;flex-shrink:0">
          <div>
            <div style="font-size:13px;font-weight:500">🔄 Reprocessar</div>
            <div style="font-size:11px;color:var(--muted)">Ignora CNPJs já consultados</div>
          </div>
        </label>
        <label class="aopt-row" style="border-color:rgba(15,122,86,.2);background:rgba(15,122,86,.03)">
          <input type="checkbox" id="aapi_so_disponiveis" onchange="toggleSoDisponiveis(this)" style="accent-color:var(--green);width:14px;height:14px;flex-shrink:0">
          <div>
            <div style="font-size:13px;font-weight:500;color:var(--green)">✅ Somente disponíveis</div>
            <div style="font-size:11px;color:var(--muted)">Reconfirma apenas CNPJs já marcados como disponível</div>
          </div>
        </label>
        <label class="aopt-row">
          <input type="checkbox" id="aapi_excluir_cliente" style="accent-color:var(--danger);width:14px;height:14px;flex-shrink:0">
          <div>
            <div style="font-size:13px;font-weight:500;color:var(--danger)">🗑 Excluir clientes do BD</div>
            <div style="font-size:11px;color:var(--muted)">Remove da tabela empresas</div>
          </div>
        </label>
        <label class="aopt-row">
          <input type="checkbox" id="aapi_extrair_clientes" style="accent-color:#1e6fa8;width:14px;height:14px;flex-shrink:0">
          <div>
            <div style="font-size:13px;font-weight:500;color:#1e6fa8">📥 Extrair clientes em Excel</div>
            <div style="font-size:11px;color:var(--muted)">Baixar lista de clientes ao final</div>
          </div>
        </label>
        <label class="aopt-row" style="border-color:rgba(180,83,9,.2);background:rgba(180,83,9,.03)">
          <input type="checkbox" id="aapi_fish" onchange="document.getElementById('fish-config').style.display=this.checked?'block':'none'" style="accent-color:#b45309;width:14px;height:14px;flex-shrink:0">
          <div>
            <div style="font-size:13px;font-weight:500;color:#b45309">🐟 Fish Mode</div>
            <div style="font-size:11px;color:var(--muted)">Envia clientes ao N8N em tempo real</div>
          </div>
        </label>
        <div id="fish-config" style="display:none;padding:10px;background:var(--surface2);border:1px solid var(--border);border-radius:6px;margin-top:2px">
          <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">
            <label class="field-label" style="margin:0">CONECTOR</label>
            <label style="display:flex;align-items:center;gap:6px;font-size:11px;color:var(--muted);cursor:pointer">
              <input type="checkbox" id="fish_split_ativo" style="accent-color:var(--accent);width:12px;height:12px" onchange="toggleFishSplit(this)">
              Dividir entre 2
            </label>
          </div>
          <div id="fish-conector-unico">
            <select id="aapi_connector_sel" onchange="toggleConectorCustom(this)" style="width:100%;margin-bottom:4px">
              <option value="">— Sem conector —</option>
              <option value="OLOS">OLOS</option>
              <option value="C6">C6</option>
              <option value="FLEX">FLEX</option>
              <option value="CRM">CRM</option>
              <option value="CUSTOM">✏ Personalizado...</option>
            </select>
            <input id="aapi_connector" type="text" placeholder="Digite o conector..." style="width:100%;display:none">
          </div>
          <div id="fish-conector-split" style="display:none;flex-direction:column;gap:8px">
            <div style="font-size:11px;color:var(--accent);padding:6px 8px;background:var(--accent-l);border-radius:5px">
              Clientes distribuídos entre os dois conectores pela % configurada
            </div>
            <div class="g2">
              <div>
                <label class="field-label">CONECTOR A</label>
                <input id="fish_connector_a" type="text" placeholder="ex: OLOS" style="width:100%;margin-bottom:4px">
                <div style="display:flex;align-items:center;gap:6px">
                  <input id="fish_pct_a" type="number" value="50" min="1" max="99" style="width:70px" oninput="sincronizarPctFish('a')">
                  <span style="font-size:12px;color:var(--muted)">%</span>
                </div>
              </div>
              <div>
                <label class="field-label">CONECTOR B</label>
                <input id="fish_connector_b" type="text" placeholder="ex: CRM" style="width:100%;margin-bottom:4px">
                <div style="display:flex;align-items:center;gap:6px">
                  <input id="fish_pct_b" type="number" value="50" min="1" max="99" style="width:70px" oninput="sincronizarPctFish('b')">
                  <span style="font-size:12px;color:var(--muted)">%</span>
                </div>
              </div>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- BOTAO INICIAR -->
  <button id="btn-aapi" onclick="iniciarAApi()" class="btn btn-primary btn-lg" style="width:100%;padding:14px;font-size:14px;letter-spacing:.5px;background:linear-gradient(135deg,#b45309,#d97706);border:none;box-shadow:0 4px 14px rgba(180,83,9,.3);transition:all .2s" onmouseover="this.style.transform='translateY(-1px)';this.style.boxShadow='0 6px 20px rgba(180,83,9,.4)'" onmouseout="this.style.transform='';this.style.boxShadow='0 4px 14px rgba(180,83,9,.3)'">
    ▶ Iniciar Consulta API C6
  </button>

  <!-- LOG PANEL -->
  <div id="aapi-log-panel" style="display:none;margin-top:20px">
    <div class="card" style="margin-bottom:10px">
      <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
        <div style="display:flex;align-items:center;gap:8px">
          <div class="section-title" style="margin:0">Controles</div>
          <span id="aapi-badge" style="padding:3px 10px;border-radius:20px;font-size:10px;font-weight:600;background:var(--accent-l);color:var(--accent)">AGUARDANDO</span>
        </div>
        <div style="display:flex;gap:6px;margin-left:4px">
          <button id="btn-pause" onclick="controleApi('pause')" class="btn btn-secondary" style="font-size:12px;padding:5px 14px">⏸ Pausar</button>
          <button id="btn-resume" onclick="controleApi('resume')" class="btn btn-secondary" style="font-size:12px;padding:5px 14px;display:none">▶ Retomar</button>
          <button onclick="controleApi('cancel')" class="btn" style="font-size:12px;padding:5px 14px;background:#fef2f2;color:var(--danger);border-color:#fecaca">🛑 Cancelar</button>
        </div>
        <div style="display:flex;align-items:center;gap:8px;margin-left:auto;flex-wrap:wrap">
          <label class="field-label" style="margin:0;white-space:nowrap;font-size:11px">Chave:</label>
          <select id="live-key" onchange="controleApi('update')" style="padding:5px 8px;border:1px solid var(--border);border-radius:6px;font-size:12px;background:var(--surface2);color:var(--text)">
            <option value="chave1">Chave 1</option>
            <option value="chave2">Chave 2</option>
            <option value="intercalar">Intercalar</option>
            <option value="dupla">Dupla</option>
          </select>
          <label class="field-label" style="margin:0;white-space:nowrap;font-size:11px">Delay (min):</label>
          <input id="live-delay" type="number" value="3" min="0" step="0.5" onchange="controleApi('update')" style="width:65px;padding:5px 8px;border:1px solid var(--border);border-radius:6px;font-size:12px;background:var(--surface2);color:var(--text)">
        </div>
      </div>
    </div>

    <!-- STATS -->
    <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:12px">
      <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:14px 16px;border-top:3px solid var(--accent)">
        <div style="font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">Consultado</div>
        <div id="aapi-stat-total" style="font-size:24px;font-weight:700;color:var(--text)">0</div>
      </div>
      <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:14px 16px;border-top:3px solid var(--green)">
        <div style="font-size:10px;color:var(--green);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">Disponível</div>
        <div id="aapi-stat-disp" style="font-size:24px;font-weight:700;color:var(--green)">0</div>
      </div>
      <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:14px 16px;border-top:3px solid var(--danger)">
        <div style="font-size:10px;color:var(--danger);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px">Cliente C6</div>
        <div id="aapi-stat-cli" style="font-size:24px;font-weight:700;color:var(--danger)">0</div>
      </div>
      <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:14px 16px;border-top:3px solid var(--warn);cursor:pointer;transition:box-shadow .15s" onclick="abrirFishDash()" onmouseover="this.style.boxShadow='0 4px 14px rgba(146,64,14,.15)'" onmouseout="this.style.boxShadow=''">
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px">
          <div style="font-size:10px;color:var(--warn);text-transform:uppercase;letter-spacing:.5px">Fish Sent</div>
          <span style="font-size:9px;color:var(--muted)">clique p/ detalhes</span>
        </div>
        <div id="aapi-stat-fish" style="font-size:24px;font-weight:700;color:var(--warn)">0</div>
      </div>
    </div>

    <!-- FISH DASH MODAL -->
    <div id="fish-dash-modal" style="display:none;position:fixed;inset:0;z-index:9000;background:rgba(0,0,0,.4);overflow:auto" onclick="if(event.target===this)this.style.display='none'">
      <div style="background:var(--surface);border-radius:var(--radius-lg);max-width:700px;margin:40px auto;padding:24px;position:relative">
        <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:20px">
          <div style="font-size:16px;font-weight:600">🐟 Fish Mode — Detalhes de Envio</div>
          <button onclick="document.getElementById('fish-dash-modal').style.display='none'" style="background:none;border:none;font-size:20px;cursor:pointer;color:var(--muted)">×</button>
        </div>
        <!-- Distribuição por conector -->
        <div id="fish-dist" style="margin-bottom:16px"></div>
        <!-- Últimos enviados -->
        <div style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);margin-bottom:8px">Últimos enviados</div>
        <div id="fish-lista" style="max-height:340px;overflow-y:auto;font-size:12px"></div>
      </div>
    </div>

    <!-- PROGRESS BAR -->
    <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:14px 16px;margin-bottom:12px">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <span style="font-size:12px;color:var(--muted)">Lote <strong id="prog-lote-atual" style="color:var(--text)">0</strong> / <span id="prog-lotes-total">?</span></span>
        <span style="font-size:12px;color:var(--muted)">Restam <strong id="prog-restam" style="color:var(--text)">—</strong> CNPJs</span>
        <span id="prog-pct" style="font-size:13px;font-weight:700;color:var(--accent)">0%</span>
      </div>
      <div style="height:8px;background:var(--surface2);border-radius:4px;overflow:hidden;border:1px solid var(--border)">
        <div id="prog-bar" style="height:100%;width:0%;background:linear-gradient(90deg,var(--accent),#7c6fe0);border-radius:4px;transition:width .6s ease"></div>
      </div>
    </div>

    <!-- LOG -->
    <div style="background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);overflow:hidden">
      <div style="padding:10px 14px;background:var(--surface2);border-bottom:1px solid var(--border);display:flex;align-items:center;gap:8px">
        <span style="width:8px;height:8px;border-radius:50%;background:var(--green);display:inline-block;animation:pulse-dot 2s infinite"></span>
        <span style="font-size:11px;font-weight:600;color:var(--muted);text-transform:uppercase;letter-spacing:.5px">Log de processamento</span>
      </div>
      <div id="aapi-log" style="padding:14px;font-family:'JetBrains Mono',monospace;font-size:11px;max-height:320px;overflow-y:auto;white-space:pre-wrap;line-height:1.7;color:var(--text)"></div>
    </div>
  </div>
</div>
<!-- TESTE FISH -->
<div id="page-teste-fish" class="page">
  <div style="margin-bottom:24px">
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:4px">
      <div style="width:36px;height:36px;border-radius:10px;background:linear-gradient(135deg,#0f7a56,#10b981);display:flex;align-items:center;justify-content:center;font-size:18px">🐟</div>
      <div>
        <div class="page-title" style="margin:0">Teste Fish Mode</div>
        <div class="page-sub" style="margin:0">Consulte um CNPJ e veja quais telefones seriam enviados ao N8N</div>
      </div>
    </div>
  </div>
  <div class="card" style="margin-bottom:16px">
    <div class="section-title">🔍 Consultar CNPJ</div>
    <div style="display:flex;gap:10px;align-items:flex-end">
      <div style="flex:1">
        <label class="field-label">CNPJ (formatado ou só números)</label>
        <input id="teste-cnpj" type="text" placeholder="ex: 11.222.333/0001-44" style="width:100%" onkeydown="if(event.key==='Enter') testarFish()">
      </div>
      <div>
        <label class="field-label">CONECTOR (opcional)</label>
        <input id="teste-connector" type="text" placeholder="ex: OLOS" style="width:130px">
      </div>
      <button onclick="testarFish()" class="btn btn-primary" style="padding:9px 24px;white-space:nowrap">🔎 Consultar</button>
    </div>
  </div>
  <div id="teste-resultado" style="display:none">
    <div class="g2" style="margin-bottom:12px">
      <div class="card">
        <div class="section-title">🏢 Dados da empresa</div>
        <div id="teste-empresa" style="font-size:13px;line-height:1.8"></div>
      </div>
      <div class="card">
        <div class="section-title">📞 Telefones encontrados</div>
        <div id="teste-fones" style="font-size:13px"></div>
      </div>
    </div>
    <div class="card" style="margin-bottom:12px">
      <div class="section-title">📤 Payload que seria enviado ao N8N</div>
      <div id="teste-payload" style="font-family:'JetBrains Mono',monospace;font-size:12px;background:var(--surface2);border:1px solid var(--border);border-radius:6px;padding:12px;white-space:pre-wrap;line-height:1.7"></div>
    </div>
    <div style="display:flex;gap:10px;align-items:center">
      <button id="btn-enviar-teste" onclick="enviarTeste()" class="btn btn-primary">🐟 Enviar para N8N agora</button>
      <span id="teste-envio-status" style="font-size:13px;color:var(--muted)"></span>
    </div>
  </div>
  <div id="teste-erro" style="display:none;padding:12px 16px;background:var(--danger-l);border:1px solid #fecaca;border-radius:8px;font-size:13px;color:var(--danger)"></div>
</div>

</div><!-- .main -->
</div><!-- .layout -->

<script>

const CAMPOS = {{ campos_json|safe }};

function goPage(id, el) {
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(n => n.classList.remove('active'));
  document.getElementById('page-' + id).classList.add('active');
  if (el) el.classList.add('active');
}
function updateToggle(el) {
  el.closest('.toggle-group').querySelectorAll('.toggle-opt').forEach(o => o.classList.toggle('on', o.querySelector('input').checked));
}
function selectFmt(el) {
  document.querySelectorAll('.fmt-radio').forEach(e => e.classList.remove('sel'));
  el.classList.add('sel'); el.querySelector('input').checked = true;
}
function toggleDualFmt(cb) {
  document.getElementById('fmt-single').style.display = cb.checked ? 'none' : 'block';
  document.getElementById('fmt-dual').style.display   = cb.checked ? 'flex'  : 'none';
  document.getElementById('dual_fmt_ativo').value     = cb.checked ? 'sim'   : 'nao';
}
function toggleModoDivisao(modo) {
  document.querySelectorAll('.div-modo-qtd').forEach(el => el.style.display = modo==='qtd' ? '' : 'none');
  document.querySelectorAll('.div-modo-pct').forEach(el => el.style.display = modo==='pct' ? '' : 'none');
  document.getElementById('info-pct').style.display = modo==='pct' ? '' : 'none';
  if (modo==='pct') atualizarSomaPct();
}
function atualizarSomaPct() {
  const a = parseFloat(document.querySelector('input[name="pct_dual_a"]')?.value)||0;
  const b = parseFloat(document.querySelector('input[name="pct_dual_b"]')?.value)||0;
  const aviso = document.getElementById('aviso-pct');
  const soma = a + b;
  if (soma !== 100 && (a>0 || b>0)) {
    aviso.style.display='block';
    aviso.textContent = `⚠ A soma das porcentagens é ${soma}% — o ideal é 100%. Serão usados os valores informados mesmo assim.`;
  } else {
    aviso.style.display='none';
  }
}
function toggleLimpeza(cb) {
  document.getElementById('limpeza-opts').style.display = cb.checked ? 'block' : 'none';
  document.getElementById('limpeza-info').style.display = cb.checked ? 'none' : 'block';
}
function updateLimpezaOpt(cb, optId) {
  const el = document.getElementById(optId);
  if (cb.checked) {
    el.style.borderColor = 'var(--accent)';
    el.style.background = 'var(--accent-l)';
  } else {
    el.style.borderColor = 'var(--border)';
    el.style.background = '';
  }
}
function limpezaPreset(tipo) {
  const opts = {
    completo:  {raiz:true,  enrich:true,  fixos:true,  blocklist:true},
    blocklist: {raiz:false, enrich:false, fixos:false, blocklist:true},
    enrich:    {raiz:false, enrich:true,  fixos:true,  blocklist:false},
    nenhum:    {raiz:false, enrich:false, fixos:false, blocklist:false},
  };
  const p = opts[tipo] || opts.nenhum;
  ['raiz','enrich','fixos','blocklist'].forEach(k => {
    const cb = document.getElementById('limpeza_'+k);
    if (cb) { cb.checked = p[k]; updateLimpezaOpt(cb, 'opt-'+k); }
  });
}
// ── Presets de exportação ──────────────────────────────────
function _lerFormExport() {
  const frm = document.getElementById('form-export');
  if (!frm) return {};
  const fd = new FormData(frm);
  const obj = {};
  for (const [k,v] of fd.entries()) {
    if (obj[k]) { if (!Array.isArray(obj[k])) obj[k]=[obj[k]]; obj[k].push(v); }
    else obj[k] = v;
  }
  // Checkboxes não marcados não aparecem no FormData — captura explicitamente
  ['split_ativo','limpeza_ativo','limpeza_raiz','limpeza_enrich','limpeza_fixos','limpeza_blocklist'].forEach(k => {
    if (!obj[k]) obj[k] = 'nao';
  });
  // Dual format
  const dualCb = document.getElementById('dual-fmt-toggle');
  obj._dual_fmt = dualCb?.checked ? 'sim' : 'nao';
  if (dualCb?.checked) {
    ['a','b'].forEach(slot => {
      obj[`_nome_dual_${slot}`]    = document.querySelector(`input[name="nome_dual_${slot}"]`)?.value || '';
      obj[`_qtd_dual_${slot}`]     = document.querySelector(`input[name="qtd_dual_${slot}"]`)?.value || '';
      obj[`_pct_dual_${slot}`]     = document.querySelector(`input[name="pct_dual_${slot}"]`)?.value || '';
      obj[`_ordem_dual_${slot}`]   = document.querySelector(`input[name="ordem_dual_${slot}"]:checked`)?.value || 'recentes';
      obj[`_formato_dual_${slot}`] = document.querySelector(`select[name="formato_dual_${slot}"]`)?.value || '';
    });
    obj._modo_divisao = document.querySelector('input[name="modo_divisao"]:checked')?.value || 'qtd';
  }
  return obj;
}
function _aplicarPresetExport(cfg) {
  // Campos simples
  ['ano_de','ano_ate','cnaes','limite','split_qtd','batch_excel','nome_lista'].forEach(k => {
    const el = document.querySelector(`[name="${k}"]`);
    if (el && cfg[k] !== undefined) el.value = cfg[k];
  });
  // Radios
  ['com_telefone','com_email','sem_mei','so_disponiveis','formato','modo_divisao'].forEach(k => {
    if (cfg[k]) {
      const el = document.querySelector(`input[name="${k}"][value="${cfg[k]}"]`);
      if (el) { el.checked = true; el.dispatchEvent(new Event('change')); }
    }
  });
  // Naturezas
  if (cfg.naturezas) {
    const nats = Array.isArray(cfg.naturezas) ? cfg.naturezas : [cfg.naturezas];
    document.querySelectorAll('input[name="naturezas"]').forEach(cb => {
      cb.checked = nats.includes(cb.value);
    });
  }
  // UFs
  if (cfg.uf) {
    const ufs = Array.isArray(cfg.uf) ? cfg.uf : [cfg.uf];
    document.querySelectorAll('input[name="uf"]').forEach(cb => {
      cb.checked = ufs.includes(cb.value);
    });
    const todos = ufs.length === document.querySelectorAll('input[name="uf"]').length;
    const ufTudo = document.getElementById('uf-tudo');
    if (ufTudo) { ufTudo.checked = todos; toggleUFPanel(ufTudo); }
  }
  // Checkboxes
  const splitCb = document.getElementById('split-toggle');
  if (splitCb) { splitCb.checked = cfg.split_ativo === 'sim'; toggleSplit(); atualizarPreviewNome(); }
  const limpCb = document.getElementById('limpeza_ativo');
  if (limpCb) { limpCb.checked = cfg.limpeza_ativo === 'sim'; toggleLimpeza(limpCb); }
  ['raiz','enrich','fixos','blocklist'].forEach(k => {
    const cb = document.getElementById('limpeza_'+k);
    if (cb) { cb.checked = cfg['limpeza_'+k] === 'sim'; updateLimpezaOpt(cb,'opt-'+k); }
  });
  // Dual format
  const dualCb = document.getElementById('dual-fmt-toggle');
  if (dualCb) {
    dualCb.checked = cfg._dual_fmt === 'sim';
    toggleDualFmt(dualCb);
    if (cfg._dual_fmt === 'sim') {
      ['a','b'].forEach(slot => {
        const n = document.querySelector(`input[name="nome_dual_${slot}"]`);
        if (n) n.value = cfg[`_nome_dual_${slot}`] || '';
        const q = document.querySelector(`input[name="qtd_dual_${slot}"]`);
        if (q) q.value = cfg[`_qtd_dual_${slot}`] || '';
        const p = document.querySelector(`input[name="pct_dual_${slot}"]`);
        if (p) p.value = cfg[`_pct_dual_${slot}`] || '';
        const ord = document.querySelector(`input[name="ordem_dual_${slot}"][value="${cfg['_ordem_dual_'+slot]||'recentes'}"]`);
        if (ord) { ord.checked = true; ord.dispatchEvent(new Event('change')); }
        const sel = document.querySelector(`select[name="formato_dual_${slot}"]`);
        if (sel && cfg[`_formato_dual_${slot}`]) sel.value = cfg[`_formato_dual_${slot}`];
      });
      if (cfg._modo_divisao) {
        const mrd = document.querySelector(`input[name="modo_divisao"][value="${cfg._modo_divisao}"]`);
        if (mrd) { mrd.checked = true; toggleModoDivisao(cfg._modo_divisao); }
      }
    }
  }
}
function salvarPreset() {
  const nome = prompt('Nome do preset:');
  if (!nome || !nome.trim()) return;
  const cfg = _lerFormExport();
  cfg._nome = nome.trim();
  const presets = JSON.parse(localStorage.getItem('exportPresets') || '[]');
  presets.push(cfg);
  localStorage.setItem('exportPresets', JSON.stringify(presets));
  renderPresets();
}
function renderPresets() {
  const list = document.getElementById('preset-list');
  if (!list) return;
  const presets = JSON.parse(localStorage.getItem('exportPresets') || '[]');
  if (!presets.length) {
    list.innerHTML = '<span style="font-size:11px;color:var(--muted)">Nenhum preset salvo</span>';
    return;
  }
  list.innerHTML = presets.map((p,i) =>
    `<div style="display:flex;align-items:center;gap:3px">
      <button type="button" onclick="carregarPreset(${i})" style="padding:4px 10px;font-size:11px;border:1px solid var(--border);border-radius:5px;background:var(--surface);cursor:pointer;color:var(--text);font-family:inherit" title="Clique para aplicar">${p._nome||'Preset '+(i+1)}</button>
      <button type="button" onclick="deletarPreset(${i})" style="padding:2px 6px;font-size:11px;border:1px solid var(--border);border-radius:5px;background:none;cursor:pointer;color:var(--muted)" title="Deletar">×</button>
    </div>`
  ).join('');
}
function carregarPreset(i) {
  const presets = JSON.parse(localStorage.getItem('exportPresets') || '[]');
  if (presets[i]) _aplicarPresetExport(presets[i]);
}
function deletarPreset(i) {
  const presets = JSON.parse(localStorage.getItem('exportPresets') || '[]');
  presets.splice(i,1);
  localStorage.setItem('exportPresets', JSON.stringify(presets));
  renderPresets();
}
// Carrega presets ao iniciar
document.addEventListener('DOMContentLoaded', renderPresets);

function toggleSplit() {
  const on = document.getElementById('split-toggle').checked;
  document.getElementById('split-options').style.display = on ? 'block' : 'none';
  document.getElementById('split-auto-info').style.display = on ? 'none' : 'block';
}
function atualizarPreviewNome() {
  const nome = document.getElementById('nome_lista').value.trim();
  const split = document.getElementById('split-toggle').checked;
  const prev  = document.getElementById('preview-nome');
  if (!nome) { prev.style.display='none'; return; }
  const base = nome.replace(/\.xlsx$/i,'');
  if (split) {
    prev.innerHTML = `📁 <strong>${base}_parte1.xlsx</strong>, <strong>${base}_parte2.xlsx</strong>... (dentro de ${base}.zip)`;
  } else {
    prev.innerHTML = `📄 <strong>${base}.xlsx</strong>`;
  }
  prev.style.display = 'block';
}
function toggleAll(v) { document.querySelectorAll('input[name=naturezas]').forEach(c => c.checked = v); }
function toggleGroup(codes) { document.querySelectorAll('input[name=naturezas]').forEach(c => { if (codes.includes(c.value)) c.checked = !c.checked; }); }
function toggleUFPanel(cb) {
  document.getElementById('uf-panel').style.display = cb.checked ? 'none' : 'block';
  if (!cb.checked) toggleUFs(true);
}
function toggleUFs(v) { document.querySelectorAll('input[name=uf]').forEach(c => { c.checked = v; updateUFStyle(c); }); }
function toggleUFGroup(ufs) { document.querySelectorAll('input[name=uf]').forEach(c => { if (ufs.includes(c.value)) { c.checked = !c.checked; updateUFStyle(c); } }); }
function updateUFStyle(cb) {
  const lbl = cb.closest('label'); if (!lbl) return;
  lbl.style.borderColor = cb.checked ? 'var(--accent)' : 'var(--border)';
  lbl.style.background  = cb.checked ? 'var(--accent-l)' : 'var(--surface2)';
  lbl.style.color       = cb.checked ? 'var(--accent)' : '';
}

// ── Format builder ─────────────────────────────────────────
let colCount = 0;
function addCol(header='', campo='razao_social', valorManual='') {
  colCount++;
  const div = document.getElementById('col-builder');
  const row = document.createElement('div'); row.className = 'col-row';
  const isManual = campo === 'manual' || valorManual !== '';
  row.innerHTML = `<span class="col-num">${colCount}</span>
    <input type="text" name="col_header[]" placeholder="Cabeçalho" value="${header}" required style="flex:1">
    <select name="col_campo[]" onchange="toggleManual(this)" style="flex:1.2">
      <option value="manual" ${isManual?'selected':''}>✏ Valor fixo manual</option>
      <option disabled>─────────────</option>
      ${Object.entries(CAMPOS).map(([k,v])=>`<option value="${k}"${k===campo&&!isManual?' selected':''}>${v}</option>`).join('')}
    </select>
    <input type="text" name="col_manual[]" placeholder="Valor fixo..." value="${valorManual}" style="flex:0.8;${isManual?'':'display:none'}">
    <button type="button" class="col-del" onclick="this.closest('.col-row').remove();renumerar()">×</button>`;
  div.appendChild(row);
}
function toggleManual(sel) {
  const inp = sel.closest('.col-row').querySelector('input[name="col_manual[]"]');
  inp.style.display = sel.value === 'manual' ? '' : 'none';
}
function renumerar() {
  document.querySelectorAll('.col-num').forEach((el,i) => el.textContent = i+1);
  colCount = document.querySelectorAll('.col-row').length;
}
function carregarFormato(key, nome, colunas) {
  document.getElementById('titulo-novo').textContent = 'Editar Formato';
  document.getElementById('formato_key').value = key;
  document.getElementById('formato_nome').value = nome;
  const idInput = document.getElementById('formato_id');
  idInput.value = key; idInput.readOnly = true;
  document.getElementById('col-builder').innerHTML = ''; colCount = 0;
  colunas.forEach(c => addCol(c.header, c.campo||'manual', c.valor_manual||''));
}
if (!document.querySelector('.col-row')) addCol();

// ── Exportação ─────────────────────────────────────────────
let exportJobId = null, exportPolling = null;
async function iniciarExportacao() {
  const form = document.getElementById('form-export');
  const data = new FormData(form);
  const panel = document.getElementById('log-panel');
  const box   = document.getElementById('log-box');
  const status= document.getElementById('log-status');
  panel.style.display = 'block'; box.textContent = '';
  document.getElementById('btn-exportar').disabled = true;
  document.getElementById('download-btn-area').innerHTML = '';
  status.textContent = 'Iniciando...'; status.style.color = '';
  const resp = await fetch('/exportar', {method:'POST', body: data});
  const json = await resp.json();
  exportJobId = json.job_id;
  exportPolling = setInterval(async () => {
    const r = await fetch('/status/' + exportJobId);
    const s = await r.json();
    box.textContent = s.log.join('\n');
    box.scrollTop = box.scrollHeight;
    if (s.done) {
      clearInterval(exportPolling);
      document.getElementById('btn-exportar').disabled = false;
      if (s.error) { status.textContent = '❌ Erro'; status.style.color='#be123c'; }
      else {
        status.textContent = '✅ ' + s.name; status.style.color = '#0f7a56';
        const a = document.createElement('a');
        a.href = '/download/' + exportJobId;
        a.textContent = '⬇ Baixar arquivo';
        a.className = 'btn btn-primary';
        a.style.fontSize = '12px'; a.style.padding = '6px 14px';
        document.getElementById('download-btn-area').appendChild(a);
      }
    }
  }, 1500);
}

// ── API Overlay ─────────────────────────────────────────────
let apiJobId=null, apiPolling=null, matrixAnim=null, clockInt=null, currentPreset='ativos';
const PRESETS = {
  recentes:   {ano_de:new Date().getFullYear()-2,ano_ate:'',limite:'',cnaes:'',naturezas:null},
  ativos:     {ano_de:'',ano_ate:'',limite:'',cnaes:'',naturezas:null},
  ei_recente: {ano_de:new Date().getFullYear()-3,ano_ate:'',limite:'',cnaes:'',naturezas:['2135','2305','2313']},
  ltda:       {ano_de:'',ano_ate:'',limite:'',cnaes:'',naturezas:['2062','2232','2240','2259','2267','2070','2089']},
  custom:     null,
};
function aplicarPreset(key) {
  // Clicou no mesmo preset ativo → desmarca
  if (currentPreset === key && key !== 'custom') {
    currentPreset = null;
    document.querySelectorAll('.preset-card, .preset-btn').forEach(b => b.classList.remove('active-preset'));
    document.getElementById('aapi-filtros-avancados').style.display = 'none';
    return;
  }
  currentPreset = key;
  document.querySelectorAll('.preset-card, .preset-btn').forEach(b => b.classList.remove('active-preset'));
  document.getElementById('preset-'+key)?.classList.add('active-preset');
  document.getElementById('aapi-filtros-avancados').style.display = key==='custom' ? 'block' : 'none';
  if (key!=='custom' && PRESETS[key]) {
    const p = PRESETS[key];
    if (document.getElementById('aapi_ano_de'))  document.getElementById('aapi_ano_de').value  = p.ano_de  || '';
    if (document.getElementById('aapi_ano_ate')) document.getElementById('aapi_ano_ate').value = p.ano_ate || '';
    if (document.getElementById('aapi_limite'))  document.getElementById('aapi_limite').value  = '';
    if (document.getElementById('aapi_cnaes'))   document.getElementById('aapi_cnaes').value   = p.cnaes   || '';
    if (p.naturezas) document.querySelectorAll('.aapi-nat').forEach(c => c.checked = p.naturezas.includes(c.value));
    else toggleAApiNat(true);
  }
}
function toggleAApiNat(v) { document.querySelectorAll('.aapi-nat').forEach(c => c.checked = v); }
function toggleAApiNatGrp2(codes) { document.querySelectorAll('.aapi-nat').forEach(c => { if (codes.includes(c.value)) c.checked = !c.checked; }); }
function toggleSoDisponiveis(cb) {
  // Somente disponíveis implica reprocessar (já estão na tabela)
  const reprocessar = document.getElementById('aapi_reprocessar');
  if (cb.checked) {
    reprocessar.checked = true;
    reprocessar.disabled = true;
    reprocessar.closest('label').style.opacity = '.5';
  } else {
    reprocessar.disabled = false;
    reprocessar.closest('label').style.opacity = '1';
  }
}
function toggleFishSplit(cb) {
  document.getElementById('fish-conector-unico').style.display = cb.checked ? 'none' : 'block';
  document.getElementById('fish-conector-split').style.display  = cb.checked ? 'flex'  : 'none';
}
function sincronizarPctFish(origem) {
  const a = document.getElementById('fish_pct_a');
  const b = document.getElementById('fish_pct_b');
  if (origem === 'a') b.value = Math.max(1, 100 - parseInt(a.value||50));
  else                a.value = Math.max(1, 100 - parseInt(b.value||50));
}
function toggleConectorCustom(sel) {
  const inp = document.getElementById('aapi_connector');
  if (sel.value==='CUSTOM') { inp.style.display=''; inp.focus(); }
  else { inp.style.display='none'; inp.value=''; }
}
function getConector() {
  const sel = document.getElementById('aapi_connector_sel');
  if (!sel) return '';
  return sel.value==='CUSTOM' ? document.getElementById('aapi_connector').value.trim() : sel.value;
}
function entrarModoApi(navBtn) {
  goPage('api', navBtn);
  // Garante que limite está vazio (sem limite = toda a base)
  const lim = document.getElementById('aapi_limite');
  if (lim && !lim.dataset.userSet) lim.value = '';
}
function entrarModoApiResultados(navBtn) {
  goPage('api-resultados', navBtn);
  // Não carrega automaticamente — usuário clica em Buscar
}
function sairModoApi() {
  const ov = document.getElementById('api-overlay');
  ov.style.opacity='0'; ov.style.transition='opacity .25s';
  setTimeout(() => { ov.style.display='none'; ov.style.pointerEvents='none'; ov.style.opacity=''; ov.style.transition=''; ov.classList.remove('show'); }, 250);
  if (matrixAnim) { cancelAnimationFrame(matrixAnim); matrixAnim=null; }
  if (clockInt)   { clearInterval(clockInt); clockInt=null; }
  document.removeEventListener('keydown', escHandler);
}
function escHandler(e) { if (e.key==='Escape') sairModoApi(); }
function iniciarMatrix() {
  const canvas = document.getElementById('api-matrix');
  const ctx = canvas.getContext('2d');
  canvas.width=window.innerWidth; canvas.height=window.innerHeight;
  const cols=Math.floor(canvas.width/15), drops=Array(cols).fill(1), chars='01ABCDEFアイウカキ9876';
  function draw() {
    ctx.fillStyle='rgba(0,0,0,.055)'; ctx.fillRect(0,0,canvas.width,canvas.height);
    ctx.font='12px monospace';
    drops.forEach((y,i) => {
      ctx.fillStyle=i%5===0?'rgba(59,130,246,.9)':'rgba(59,130,246,.4)';
      ctx.fillText(chars[Math.floor(Math.random()*chars.length)],i*15,y*15);
      if(y*15>canvas.height&&Math.random()>.975) drops[i]=0; drops[i]++;
    });
    matrixAnim=requestAnimationFrame(draw);
  }
  draw();
}
function aApiLog(msg, color) {
  const box=document.getElementById('aapi-log');
  const line=document.createElement('div'); line.style.color=color||'#4f8ef7'; line.textContent=msg;
  box.appendChild(line); box.scrollTop=box.scrollHeight;
}
function atualizarInfoLote(sel) {
  const isDupla = sel.value === 'dupla';
  const lote = isDupla ? 40000 : 20000;
  document.getElementById('info-lote').textContent = `Lote: ${lote.toLocaleString('pt-BR')} CNPJs por execução${isDupla ? ' (20k por chave)' : ''}`;
}

async function controleApi(action) {
  if (!apiJobId) { console.warn('No active job'); return; }
  const body = { action };
  if (action === 'update') {
    body.key_mode  = document.getElementById('live-key').value;
    body.delay_min = document.getElementById('live-delay').value;
    body.batch_size = body.key_mode === 'dupla' ? 40000 : 20000;
  }
  try {
    const r = await fetch(`/api/controle/${apiJobId}`, {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(body)});
    await r.json();
  } catch(e) { console.error('controleApi error:', e); }
  if (action === 'pause') {
    document.getElementById('btn-pause').style.display  = 'none';
    document.getElementById('btn-resume').style.display = '';
  } else if (action === 'resume') {
    document.getElementById('btn-pause').style.display  = '';
    document.getElementById('btn-resume').style.display = 'none';
  } else if (action === 'cancel') {
    if (apiPolling) { clearInterval(apiPolling); apiPolling = null; }
    document.getElementById('btn-aapi').disabled = false;
    document.getElementById('aapi-badge').textContent = 'CANCELADO';
    document.getElementById('aapi-badge').style.color = 'var(--danger)';
    document.getElementById('btn-pause').style.display  = '';
    document.getElementById('btn-resume').style.display = 'none';
    apiJobId = null;
  }
}

async function iniciarAApi() {
  const naturezas = [...document.querySelectorAll('.aapi-nat:checked')].map(c => c.value);
  const keyMode   = document.getElementById('aapi_key_mode').value;
  const batchSize = keyMode === 'dupla' ? 40000 : 20000;
  const body = {
    preset:          currentPreset,
    ano_de:          document.getElementById('aapi_ano_de').value,
    ano_ate:         document.getElementById('aapi_ano_ate').value,
    limite:          0,  // sem limite — processa toda a base
    cnaes:           document.getElementById('aapi_cnaes').value,
    sem_mei:         'sim',
    key_mode:        keyMode,
    batch_size:      batchSize,
    delay_min:       document.getElementById('aapi_delay').value,
    naturezas,
    fish_mode:       document.getElementById('aapi_fish').checked,
    connector:       getConector(),
    fish_split:       document.getElementById('fish_split_ativo')?.checked || false,
    fish_connector_a: document.getElementById('fish_connector_a')?.value.trim() || '',
    fish_connector_b: document.getElementById('fish_connector_b')?.value.trim() || '',
    fish_pct_a:       parseInt(document.getElementById('fish_pct_a')?.value||50),
    reprocessar:          document.getElementById('aapi_reprocessar').checked,
    so_disponiveis_api:   document.getElementById('aapi_so_disponiveis').checked,
    excluir_cliente:      document.getElementById('aapi_excluir_cliente').checked,
    extrair_clientes:document.getElementById('aapi_extrair_clientes').checked,
  };
  document.getElementById('aapi-log-panel').style.display='block';
  document.getElementById('aapi-log').innerHTML='';
  // Sync live controls with initial values
  document.getElementById('live-key').value   = body.key_mode;
  document.getElementById('live-delay').value = body.delay_min;
  document.getElementById('btn-pause').style.display  = '';
  document.getElementById('btn-resume').style.display = 'none';
  document.getElementById('aapi-badge').textContent='PROCESSANDO'; document.getElementById('aapi-badge').style.color='#4f8ef7';
  ['aapi-stat-total','aapi-stat-disp','aapi-stat-cli','aapi-stat-fish'].forEach(id => document.getElementById(id).textContent='0');
  document.getElementById('btn-aapi').disabled=true;
  
  if (body.fish_mode) {
    if (body.fish_split && body.fish_connector_a && body.fish_connector_b) {
      aApiLog(`> FISH MODE ativo — ${body.fish_pct_a}% → ${body.fish_connector_a} | ${100-body.fish_pct_a}% → ${body.fish_connector_b}`,'#ffaa22');
    } else {
      aApiLog('> FISH MODE ativo — conector: '+(body.connector||'nenhum'),'#ffaa22');
    }
  }
  try {
    const resp=await fetch('/api/iniciar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)});

    const json=await resp.json(); apiJobId=json.job_id;
  } catch(e) { aApiLog('ERRO: '+e.message,'#ff5555'); document.getElementById('btn-aapi').disabled=false; return; }
  let lastLen=0;
  apiPolling=setInterval(async()=>{
    try {
      const r=await fetch('/api/status/'+apiJobId); const s=await r.json();
      s.log.slice(lastLen).forEach(line=>{
        const c=line.includes('❌')||line.includes('ERRO')?'#ff5555':line.includes('disponivel')?'#5090ff':line.includes('cliente')?'#ff7744':line.includes('FISH')?'#ffaa22':'#4f8ef7';
        aApiLog(line,c);
      });
      lastLen=s.log.length;
      if (s.stats) {
        document.getElementById('aapi-stat-total').textContent=s.stats.total.toLocaleString('pt-BR');
        document.getElementById('aapi-stat-disp').textContent=s.stats.disp.toLocaleString('pt-BR');
        document.getElementById('aapi-stat-cli').textContent=s.stats.cli.toLocaleString('pt-BR');
        document.getElementById('aapi-stat-fish').textContent=s.stats.fish.toLocaleString('pt-BR');
        // Progress bar
        const loteAtual  = s.stats.lote_atual  || 0;
        const lotesTotal = s.stats.lotes_total || 0;
        const restam     = s.stats.restam_cnpjs != null ? s.stats.restam_cnpjs : null;
        const pct        = lotesTotal > 0 ? Math.round((loteAtual / lotesTotal) * 100) : 0;
        document.getElementById('prog-lote-atual').textContent = loteAtual.toLocaleString('pt-BR');
        document.getElementById('prog-lotes-total').textContent = lotesTotal.toLocaleString('pt-BR');
        document.getElementById('prog-restam').textContent = restam != null ? restam.toLocaleString('pt-BR') : '—';
        document.getElementById('prog-pct').textContent = pct + '%';
        document.getElementById('prog-bar').style.width = pct + '%';
      }
      if (s.done) {
        clearInterval(apiPolling); document.getElementById('btn-aapi').disabled=false;
        document.getElementById('aapi-badge').textContent=s.error?'ERRO':'CONCLUIDO';
        document.getElementById('aapi-badge').style.color=s.error?'#ff5555':'#4f8ef7';
        if (s.has_clientes_xlsx) {
          const box=document.getElementById('aapi-log');
          const btn=document.createElement('a');
          btn.href=`/api/download-clientes/${apiJobId}`;
          btn.textContent=`⬇ BAIXAR CLIENTES (${s.clientes_count.toLocaleString()} registros)`;
          btn.style.cssText='display:inline-block;margin-top:10px;padding:8px 16px;background:rgba(80,144,255,.1);border:1px solid rgba(80,144,255,.5);color:#5090ff;font-family:inherit;font-size:10px;border-radius:3px;text-decoration:none;letter-spacing:1px';
          box.appendChild(btn); box.scrollTop=box.scrollHeight;
        }
      }
    } catch(e) {}
  },2000);
}

// ── Dashboard ───────────────────────────────────────────────
async function carregarResultados() {
  const filtro=document.getElementById('res-filtro').value;
  const limite=document.getElementById('res-limite').value;
  const r=await fetch(`/api/resultados?resultado=${filtro}&limite=${limite}`);
  const data=await r.json();
  const total=data.total||0, disp=data.disponiveis||0, cli=data.clientes||0, totalBd=data.total_bd||0;
  const naoProc=totalBd-total;
  const pctProc=totalBd?((total/totalBd)*100).toFixed(1):0;
  const pctDisp=total?((disp/total)*100).toFixed(1):0;
  const pctCli=total?((cli/total)*100).toFixed(1):0;
  document.getElementById('dc-total-bd').textContent=totalBd.toLocaleString('pt-BR');
  document.getElementById('dc-processados').textContent=total.toLocaleString('pt-BR');
  document.getElementById('dc-processados-pct').textContent=`${pctProc}% da base`;
  document.getElementById('dc-disp').textContent=disp.toLocaleString('pt-BR');
  document.getElementById('dc-disp-pct').textContent=`${pctDisp}% dos processados`;
  document.getElementById('dc-cli').textContent=cli.toLocaleString('pt-BR');
  document.getElementById('dc-cli-pct').textContent=`${pctCli}% dos processados`;
  drawDonut('chart-cobertura','legend-cobertura',[
    {label:'Processados',value:total,color:'#5b4fcf'},
    {label:'Não processados',value:naoProc,color:'#e0dcd4'},
  ]);
  drawDonut('chart-resultado','legend-resultado',[
    {label:'Disponível',value:disp,color:'#0f7a56'},
    {label:'Cliente C6',value:cli,color:'#be123c'},
  ]);
  const barsEl=document.getElementById('dash-bars');
  const bars=[
    {label:'Cobertura da base',value:pctProc,color:'#5b4fcf',fmt:`${total.toLocaleString('pt-BR')} / ${totalBd.toLocaleString('pt-BR')}`},
    {label:'Taxa de disponibilidade',value:pctDisp,color:'#0f7a56',fmt:`${disp.toLocaleString('pt-BR')} disponíveis`},
    {label:'Taxa de clientes C6',value:pctCli,color:'#be123c',fmt:`${cli.toLocaleString('pt-BR')} clientes`},
  ];
  barsEl.innerHTML=bars.map(b=>`<div class="bar-row"><div class="bar-label"><span>${b.label}</span><span style="font-weight:500;color:var(--text)">${b.value}% <span style="color:var(--muted);font-weight:400">${b.fmt}</span></span></div><div class="bar-track"><div class="bar-fill" style="width:0%;background:${b.color}" data-target="${b.value}"></div></div></div>`).join('');
  requestAnimationFrame(()=>{ document.querySelectorAll('.bar-fill').forEach(el=>{ el.style.width=Math.min(100,el.dataset.target)+'%'; }); });

  // ── Naturezas breakdown ────────────────────────────────
  const nat = data.naturezas_stats || [];
  if (nat.length) {
    const pctF = (a,b) => b > 0 ? ((a/b)*100).toFixed(1)+'%' : '—';
    let ht = `<table style="width:100%;border-collapse:collapse;font-size:12px">
      <thead><tr style="background:var(--surface2)">
        <th style="text-align:left;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--muted);font-size:11px;text-transform:uppercase">Natureza</th>
        <th style="text-align:left;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--muted);font-size:11px;text-transform:uppercase">Cód</th>
        <th style="text-align:right;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--muted);font-size:11px;text-transform:uppercase">Total</th>
        <th style="text-align:right;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--muted);font-size:11px;text-transform:uppercase">Processados</th>
        <th style="text-align:right;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--green);font-size:11px;text-transform:uppercase">Disponível</th>
        <th style="text-align:right;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--danger);font-size:11px;text-transform:uppercase">Cliente</th>
        <th style="text-align:left;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--muted);font-size:11px;text-transform:uppercase">Cobertura</th>
        <th style="text-align:left;padding:7px 10px;border-bottom:2px solid var(--border);color:var(--muted);font-size:11px;text-transform:uppercase">Conversão</th>
      </tr></thead><tbody>`;
    nat.forEach((r,i) => {
      const bg = i%2===0?'':'background:var(--surface2)';
      const cob = r.total > 0 ? ((r.processados/r.total)*100).toFixed(1) : 0;
      const conv = r.processados > 0 ? ((r.disponiveis/r.processados)*100).toFixed(1) : 0;
      ht += `<tr style="${bg}">
        <td style="padding:6px 10px;border-bottom:1px solid var(--border)">${r.nome||'—'}</td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border);font-family:monospace;color:var(--green)">${r.cod}</td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border);text-align:right">${Number(r.total).toLocaleString('pt-BR')}</td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border);text-align:right">${Number(r.processados).toLocaleString('pt-BR')}</td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border);text-align:right;color:var(--green);font-weight:500">${Number(r.disponiveis).toLocaleString('pt-BR')}</td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border);text-align:right;color:var(--danger);font-weight:500">${Number(r.clientes).toLocaleString('pt-BR')}</td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border)">
          <div style="display:flex;align-items:center;gap:6px">
            <div style="flex:1;height:6px;background:var(--surface2);border-radius:3px;overflow:hidden;border:1px solid var(--border)"><div style="height:100%;width:${Math.min(100,cob)}%;background:var(--accent);border-radius:3px;transition:width .6s ease"></div></div>
            <span style="font-size:11px;color:var(--muted);white-space:nowrap">${cob}%</span>
          </div>
        </td>
        <td style="padding:6px 10px;border-bottom:1px solid var(--border)">
          <div style="display:flex;align-items:center;gap:6px">
            <div style="flex:1;height:6px;background:var(--surface2);border-radius:3px;overflow:hidden;border:1px solid var(--border)"><div style="height:100%;width:${Math.min(100,conv)}%;background:var(--green);border-radius:3px;transition:width .6s ease"></div></div>
            <span style="font-size:11px;color:var(--muted);white-space:nowrap">${conv}%</span>
          </div>
        </td>
      </tr>`;
    });
    ht += '</tbody></table>';
    document.getElementById('dash-naturezas').innerHTML = ht;
  }
  if (!data.rows.length) { document.getElementById('res-tabela').innerHTML='<p style="color:var(--muted);font-size:13px;padding:16px 0">Nenhum resultado.</p>'; return; }
  const cols=['cnpj','status','processado_em'];
  let html='<table style="width:100%;border-collapse:collapse"><tr>'+cols.map(c=>`<th style="text-align:left;padding:7px 12px;border-bottom:2px solid var(--border);font-size:11px;text-transform:uppercase;color:var(--muted)">${c}</th>`).join('')+'</tr>';
  data.rows.forEach((row,i)=>{
    const cor=row.status==='disponivel'?'var(--green)':'var(--danger)';
    const bg=i%2===0?'':'background:var(--surface2)';
    html+=`<tr style="${bg}">` + cols.map(c => {
      if (c === 'status') {
        const bg2 = row.status==='disponivel' ? 'var(--green-l)' : 'var(--danger-l)';
        const fg  = row.status==='disponivel' ? 'var(--green)'   : 'var(--danger)';
        return `<td style="padding:6px 12px;border-bottom:1px solid var(--border)"><span style="padding:2px 8px;border-radius:20px;font-size:11px;font-weight:500;background:${bg2};color:${fg}">${row[c]||''}</span></td>`;
      }
      return `<td style="padding:6px 12px;border-bottom:1px solid var(--border);color:var(--text)">${row[c]||''}</td>`;
    }).join('') + '</tr>';
  });
  document.getElementById('res-tabela').innerHTML=html+'</table>';
}
function drawDonut(canvasId,legendId,segments) {
  const canvas=document.getElementById(canvasId); if(!canvas) return;
  const ctx=canvas.getContext('2d'), cx=canvas.width/2, cy=canvas.height/2, r=55, inner=34;
  const total=segments.reduce((s,seg)=>s+(seg.value||0),0);
  ctx.clearRect(0,0,canvas.width,canvas.height);
  if (total===0) { ctx.beginPath(); ctx.arc(cx,cy,r,0,Math.PI*2); ctx.strokeStyle='#e0dcd4'; ctx.lineWidth=r-inner; ctx.stroke(); }
  else {
    let start=-Math.PI/2;
    segments.forEach(seg=>{
      if(!seg.value) return;
      const angle=(seg.value/total)*Math.PI*2;
      ctx.beginPath(); ctx.arc(cx,cy,(r+inner)/2,start,start+angle);
      ctx.strokeStyle=seg.color; ctx.lineWidth=r-inner; ctx.stroke(); start+=angle;
    });
  }
  const pct=total>0?((segments[0].value/total)*100).toFixed(0):'0';
  ctx.fillStyle='#1c1917'; ctx.font='bold 18px Inter,sans-serif';
  ctx.textAlign='center'; ctx.textBaseline='middle'; ctx.fillText(pct+'%',cx,cy);
  document.getElementById(legendId).innerHTML=segments.map(seg=>`<div style="display:flex;align-items:center;gap:8px"><div style="width:10px;height:10px;border-radius:2px;background:${seg.color};flex-shrink:0"></div><div><div style="font-size:12px;color:var(--text);font-weight:500">${seg.value.toLocaleString('pt-BR')}</div><div style="font-size:11px;color:var(--muted)">${seg.label}</div></div></div>`).join('');
}
// ── Teste Fish ───────────────────────────────────────────────
let _testePayload = null;
async function testarFish() {
  const cnpj = document.getElementById('teste-cnpj').value.trim();
  const connector = document.getElementById('teste-connector').value.trim();
  if (!cnpj) return;
  document.getElementById('teste-resultado').style.display = 'none';
  document.getElementById('teste-erro').style.display = 'none';
  document.getElementById('teste-envio-status').textContent = '';
  const r = await fetch('/api/teste-fish', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({cnpj, connector})});
  const data = await r.json();
  if (data.erro) {
    document.getElementById('teste-erro').textContent = '❌ ' + data.erro;
    document.getElementById('teste-erro').style.display = 'block';
    return;
  }
  _testePayload = data;
  const e = data.empresa;
  const statusColor = e.status_api==='disponivel' ? 'var(--green)' : e.status_api==='cliente' ? 'var(--danger)' : 'var(--muted)';
  document.getElementById('teste-empresa').innerHTML = `
    <div><strong>Razão Social:</strong> ${e.razao_social||'—'}</div>
    <div><strong>CNPJ:</strong> ${e.cnpj||'—'}</div>
    <div><strong>Email:</strong> ${e.email||'—'}</div>
    <div><strong>Cidade/UF:</strong> ${e.cidade||'—'} / ${e.estado||'—'}</div>
    <div><strong>Status API C6:</strong> <span style="color:${statusColor};font-weight:500">${e.status_api||'não consultado'}</span></div>
  `;
  const fones = data.fones;
  const foneSrc = data.fones_source;
  if (!fones.length) {
    document.getElementById('teste-fones').innerHTML = '<span style="color:var(--danger)">❌ Nenhum telefone encontrado — não seria enviado ao N8N</span>';
    document.getElementById('btn-enviar-teste').disabled = true;
  } else {
    document.getElementById('btn-enviar-teste').disabled = false;
    document.getElementById('teste-fones').innerHTML = fones.map((f,i) =>
      `<div style="display:flex;align-items:center;gap:10px;padding:7px 0;border-bottom:1px solid var(--border)">
        <span style="font-family:'JetBrains Mono',monospace;font-size:14px;font-weight:600">${f}</span>
        <span style="font-size:11px;color:var(--muted)">fone${i+1}</span>
        ${(foneSrc==='externo')?'<span style="font-size:10px;padding:2px 8px;background:var(--warn-l);color:var(--warn);border-radius:20px;border:1px solid #fde68a">banco externo</span>':'<span style="font-size:10px;padding:2px 8px;background:var(--green-l);color:var(--green);border-radius:20px">empresa</span>'}
      </div>`
    ).join('') + `<div style="margin-top:8px;font-size:11px;color:var(--muted)">📍 Fonte: ${foneSrc==='empresa'?'tabela empresas (banco principal)':'tabela telefones (banco externo)'}</div>`;
  }
  document.getElementById('teste-payload').textContent = JSON.stringify(data.payload, null, 2);
  document.getElementById('teste-resultado').style.display = 'block';
}
async function abrirFishDash() {
  if (!apiJobId) return;
  const r = await fetch('/api/fish-log/' + apiJobId);
  const data = await r.json();
  const modal = document.getElementById('fish-dash-modal');

  // Distribuição por conector
  const dist = data.dist || {};
  const total = Object.values(dist).reduce((a,b)=>a+b,0) || 1;
  const distHtml = Object.keys(dist).length
    ? `<div style="margin-bottom:4px;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.5px;color:var(--muted)">Distribuição por conector</div>
       <div style="display:flex;flex-direction:column;gap:8px;margin-bottom:16px">` +
      Object.entries(dist).map(([conn, n]) => {
        const pct = ((n/total)*100).toFixed(1);
        return `<div>
          <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:3px">
            <strong>${conn||'sem conector'}</strong>
            <span>${n.toLocaleString('pt-BR')} <span style="color:var(--muted)">(${pct}%)</span></span>
          </div>
          <div style="height:6px;background:var(--surface2);border-radius:3px;overflow:hidden;border:1px solid var(--border)">
            <div style="height:100%;width:${pct}%;background:var(--warn);border-radius:3px;transition:width .5s"></div>
          </div>
        </div>`;
      }).join('') + '</div>'
    : '';
  document.getElementById('fish-dist').innerHTML = distHtml;

  // Últimos enviados
  const rows = data.log || [];
  if (!rows.length) {
    document.getElementById('fish-lista').innerHTML = '<p style="color:var(--muted)">Nenhum envio registrado ainda.</p>';
  } else {
    let html = '<table style="width:100%;border-collapse:collapse">';
    html += '<tr style="border-bottom:2px solid var(--border)">' +
      ['CNPJ','Razão Social','Fones','Conector'].map(h =>
        `<th style="text-align:left;padding:6px 8px;font-size:10px;color:var(--muted);text-transform:uppercase">${h}</th>`
      ).join('') + '</tr>';
    rows.slice().reverse().forEach(r => {
      html += `<tr style="border-bottom:1px solid var(--border)">
        <td style="padding:6px 8px;font-family:'JetBrains Mono',monospace;font-size:11px">${r.cnpj}</td>
        <td style="padding:6px 8px;font-size:12px;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${r.nome||'—'}</td>
        <td style="padding:6px 8px;font-size:11px;color:var(--green)">${(r.fones||[]).join(', ')||'—'}</td>
        <td style="padding:6px 8px"><span style="padding:2px 8px;border-radius:20px;font-size:10px;font-weight:500;background:var(--warn-l);color:var(--warn)">${r.conector||'—'}</span></td>
      </tr>`;
    });
    document.getElementById('fish-lista').innerHTML = html + '</table>';
  }
  modal.style.display = 'block';
}

async function enviarTeste() {
  if (!_testePayload?.payload) return;
  const btn = document.getElementById('btn-enviar-teste');
  const status = document.getElementById('teste-envio-status');
  btn.disabled = true; status.textContent = '⏳ Enviando...'; status.style.color = 'var(--muted)';
  const r = await fetch('/api/teste-fish-enviar', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(_testePayload.payload)});
  const data = await r.json();
  status.textContent = data.ok ? '✅ Enviado com sucesso!' : '❌ Erro: ' + data.erro;
  status.style.color = data.ok ? 'var(--green)' : 'var(--danger)';
  btn.disabled = false;
}

function exportarResultados() {
  window.location.href=`/api/exportar-resultados?resultado=${document.getElementById('res-filtro').value}`;
}

// init
document.addEventListener('DOMContentLoaded', () => {
  document.getElementById('aapi_fish').addEventListener('change', function() {
    document.getElementById('fish-config').style.display = this.checked ? 'flex' : 'none';
  });
});

</script>
{% if editar_key %}
<script>
document.addEventListener('DOMContentLoaded',function(){
  carregarFormato({{ editar_key|tojson }},{{ editar_nome|tojson }},{{ editar_colunas|safe }});
  goPage('novo-formato', document.querySelectorAll('.nav-btn')[2]);
});
</script>
{% endif %}
</body>
</html>"""

app = Flask(__name__)

def get_conn():
    return psycopg2.connect(
        DATABASE_URL,
        sslmode='require',
        connect_timeout=15,
        keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=5
    )

def get_ext_conn():
    """Conexão com banco externo (telefones, blocklist, raiz) com timeout e keepalives."""
    url = TELEFONES_DB_URL.replace("&channel_binding=require","")
    return psycopg2.connect(
        url,
        sslmode='require',
        connect_timeout=15,
        keepalives=1, keepalives_idle=30, keepalives_interval=10, keepalives_count=5
    )

def carregar_formatos():
    if FORMATOS_FILE.exists():
        with open(FORMATOS_FILE) as f:
            return json.load(f)
    return FORMATOS_PADRAO.copy()

def salvar_formatos(formatos):
    with open(FORMATOS_FILE, "w") as f:
        json.dump(formatos, f, indent=2, ensure_ascii=False)

def limpar_num(v):
    if not v: return None
    s = re.sub(r'\D','',str(v))
    try: return int(s) if s else None
    except: return None

def clean_name(n):
    if not n: return ""
    n = re.sub(r'\s+\d{11}$','',str(n).strip())
    n = re.sub(r'\s+\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}$','',n)
    return n.strip()

NUM_CAMPOS = {"cnpj_numerico","telefone_principal_num","telefone_secundario_num","atividade_principal_cod_num"}

def colunas_necessarias(formato):
    """Retorna apenas as colunas do BD que o formato precisa — evita SELECT *"""
    # Mapeamento: campo do formato → coluna(s) real(is) no banco
    MAPA = {
        "razao_social":               ["razao_social"],
        "cnpj":                       ["cnpj"],
        "cnpj_numerico":              ["cnpj"],
        "email":                      ["email"],
        "telefone_principal":         ["telefone_principal"],
        "telefone_principal_num":     ["telefone_principal"],
        "telefone_secundario":        ["telefone_secundario"],
        "telefone_secundario_num":    ["telefone_secundario"],
        "atividade_principal_cod":    ["atividade_principal_cod"],
        "atividade_principal_cod_num":["atividade_principal_cod"],
        "atividade_principal":        ["atividade_principal"],
        "natureza_juridica_cod":      ["natureza_juridica_cod"],
        "natureza_juridica":          ["natureza_juridica"],
        "data_abertura":              ["data_abertura"],
        "ano_abertura":               ["data_abertura"],
        "estado":                     ["estado"],
        "cidade":                     ["cidade"],
        "logradouro":                 ["logradouro"],
        "bairro":                     ["bairro"],
        "cep":                        ["cep"],
        "nome_socio":                 ["nome_socio"],
        "cpf_socio":                  ["cpf_socio"],
    }
    cols = set()
    for col_def in formato.get("colunas", []):
        campo = col_def.get("campo", "")
        for db_col in MAPA.get(campo, []):
            cols.add(db_col)
    # cnpj sempre necessário (chave)
    cols.add("cnpj")
    return sorted(cols)
    campo = col_def.get('campo', '')
    if campo == 'manual':
        return col_def.get('valor_manual', '')
    return {
        "razao_social":               clean_name(row.get("razao_social","")),
        "cnpj":                       row.get("cnpj",""),
        "cnpj_numerico":              limpar_num(row.get("cnpj")),
        "email":                      row.get("email","") or "",
        "telefone_principal":         row.get("telefone_principal","") or "",
        "telefone_principal_num":     limpar_num(row.get("telefone_principal")),
        "telefone_secundario":        row.get("telefone_secundario","") or "",
        "telefone_secundario_num":    limpar_num(row.get("telefone_secundario")),
        "atividade_principal_cod":    row.get("atividade_principal_cod","") or "",
        "atividade_principal_cod_num":limpar_num(row.get("atividade_principal_cod")),
        "atividade_principal":        row.get("atividade_principal","") or "",
        "natureza_juridica_cod":      row.get("natureza_juridica_cod","") or "",
        "natureza_juridica":          row.get("natureza_juridica","") or "",
        "data_abertura":              data_str,
        "ano_abertura":               ano_str,
        "estado":                     row.get("estado","") or "",
        "cidade":                     row.get("cidade","") or "",
        "logradouro":                 row.get("logradouro","") or "",
        "bairro":                     row.get("bairro","") or "",
        "cep":                        row.get("cep","") or "",
        "nome_socio":                 row.get("nome_socio","") or "",
        "cpf_socio":                  row.get("cpf_socio","") or "",
        "fixo_olos":"OLOS","fixo_flex":"FLEX","fixo_c6":"C6","fixo_vazio":None,
    }.get(campo,"")

BATCH_XLSX = 5000  # linhas por lote no Excel (evita travar)

def inserir9(tel):
    """
    Insere o 9 de celular após o DDD quando necessário.
    Regra: após o DDD (2 dígitos), se o próximo dígito for 6-9 = celular.
    - 10 dígitos + começa com 6-9 → insere 9 (ex: 2196299472 → 21996299472)
    - 11 dígitos → já tem 9, deixa como está
    - Fixo (começa com 2-5 após DDD) → não insere 9
    """
    n = re.sub(r'\D', '', str(tel or ''))
    if len(n) == 10:
        primeiro_apos_ddd = n[2]  # dígito logo após os 2 do DDD
        if primeiro_apos_ddd in '6789':
            return n[:2] + '9' + n[2:]  # celular sem 9 → insere
        return n  # fixo → não mexe
    return n  # 11 dígitos já correto, ou outro formato

def resolver(col_def, row, ano_str, data_str):
    campo = col_def.get('campo', '')
    if campo == 'manual':
        return col_def.get('valor_manual', '')
    return {
        "razao_social":               clean_name(row.get("razao_social","")),
        "cnpj":                       row.get("cnpj",""),
        "cnpj_numerico":              limpar_num(row.get("cnpj")),
        "email":                      row.get("email","") or "",
        "telefone_principal":         row.get("telefone_principal","") or "",
        "telefone_principal_num":     inserir9(row.get("telefone_principal","")),
        "telefone_secundario":        row.get("telefone_secundario","") or "",
        "telefone_secundario_num":    inserir9(row.get("telefone_secundario","")),
        "atividade_principal_cod":    row.get("atividade_principal_cod","") or "",
        "atividade_principal_cod_num":limpar_num(row.get("atividade_principal_cod")),
        "atividade_principal":        row.get("atividade_principal","") or "",
        "natureza_juridica_cod":      row.get("natureza_juridica_cod","") or "",
        "natureza_juridica":          row.get("natureza_juridica","") or "",
        "data_abertura":              data_str,
        "ano_abertura":               ano_str,
        "estado":                     row.get("estado","") or "",
        "cidade":                     row.get("cidade","") or "",
        "logradouro":                 row.get("logradouro","") or "",
        "bairro":                     row.get("bairro","") or "",
        "cep":                        row.get("cep","") or "",
        "nome_socio":                 row.get("nome_socio","") or "",
        "cpf_socio":                  row.get("cpf_socio","") or "",
        "fixo_olos":"OLOS","fixo_flex":"FLEX","fixo_c6":"C6","fixo_vazio":None,
    }.get(campo,"")


def fazer_xlsx(rows, formato, job_id=None):
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Leads"
    headers = [c['header'] for c in formato['colunas']]
    thin = Side(style="thin", color="e0dcd4")
    brd  = Border(bottom=thin)
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_fill = PatternFill("solid", fgColor="5b4fcf")
    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = ctr
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.row_dimensions[1].height = 22

    total = len(rows)
    for batch_start in range(0, total, BATCH_XLSX):
        batch = rows[batch_start:batch_start + BATCH_XLSX]
        if job_id:
            pct = min(100, int((batch_start / total) * 100))
            _log(job_id, f"   escrevendo linhas {batch_start+1:,}–{min(batch_start+BATCH_XLSX, total):,} ({pct}%)")
        for row in batch:
            da = row.get('data_abertura')
            ano_str  = str(da.year) if da else ""
            data_str = da.strftime('%d/%m/%Y') if da else ""
            row_data = [resolver(c, row, ano_str, data_str) for c in formato['colunas']]
            ws.append(row_data)
            r = ws.max_row
            for c_idx, (val, cd) in enumerate(zip(row_data, formato['colunas']), 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.alignment = lft
                cell.border = brd
                if cd.get('campo') in NUM_CAMPOS and val is not None:
                    cell.number_format = '0'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def filtros_default():
    return dict(ano_de='',ano_ate='',cnaes='',limite='',
                com_telefone='nao',com_email='nao',sem_mei='sim',so_disponiveis='nao',
                naturezas_sel=list(NATUREZAS.keys()),ufs_sel=[],
                formato='bernardo',split_ativo='nao',split_qtd='500000',
                batch_excel='200000',nome_lista='')

@app.route('/', methods=['GET'])
def index():
    resultado = request.args.get('resultado')
    return render_template_string(HTML,
        naturezas=NATUREZAS, ufs=UFS,
        f=filtros_default(), formatos=carregar_formatos(),
        campos_json=json.dumps(CAMPOS_DISPONIVEIS),
        resultado=resultado)

import threading
import uuid
import time

_export_status = {}   # job_id → {"log": [...], "done": bool, "file": bytes, "name": str, "error": str}

def novo_job():
    import uuid
    return str(uuid.uuid4())[:8]

@app.route('/exportar', methods=['POST'])
def exportar():
    frm = request.form
    job_id = novo_job()
    _export_status[job_id] = {"log": [], "done": False, "file": None, "name": "", "error": ""}

    # Roda em thread para não bloquear
    t = threading.Thread(target=_executar_exportacao, args=(job_id, frm))
    t.daemon = True
    t.start()

    from flask import jsonify
    return jsonify({"job_id": job_id})

def _log(job_id, msg):
    _export_status[job_id]["log"].append(msg)

def _eh_fixo(num: str) -> bool:
    """Retorna True se o número for telefone fixo (não deve ser incluído)."""
    n = re.sub(r'\D','',str(num or ''))
    if len(n) == 10:
        return int(n[2]) <= 5  # fixo: 3º dígito ≤ 5
    if len(n) == 11 and int(n[2]) == 9:
        return False  # celular com 9
    if len(n) in (8, 9):
        return True   # sem DDD = fixo
    return False

def _gerar_xlsx_parcial(cols_sql, wheres, params, limite, ordem, formato, batch_excel, nome, job_id,
                        split_qtd=None, zf_externo=None,
                        usar_limpeza=False, usar_raiz=False, usar_enrich=False,
                        usar_fixos=False, usar_blocklist=False, raiz_set=None):
    """
    Gera xlsx(s) com ORDER BY data_abertura ASC/DESC.
    Aplica pipeline de limpeza se usar_limpeza=True.
    Retorna (bytes_or_None, total_linhas).
    """
    order_dir = "DESC" if ordem == "recentes" else "ASC"
    raiz_set  = raiz_set or set()

    # Stats de limpeza
    st = {"lidos":0, "removidos_raiz":0, "enriquecidos":0, "fones_extras":0,
          "fixos_removidos":0, "bloqueados":0, "escritos":0}

    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    hdr_fill = PatternFill("solid", fgColor="5b4fcf")

    def _novo_wb():
        _wb = Workbook(write_only=True)
        _ws = _wb.create_sheet("Leads")
        from openpyxl.cell import WriteOnlyCell
        hdr_row = []
        for h in [c['header'] for c in formato['colunas']]:
            cell = WriteOnlyCell(_ws, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            hdr_row.append(cell)
        _ws.append(hdr_row)
        return _wb, _ws

    chunk    = split_qtd or limite
    total    = 0
    offset   = 0
    file_idx = 1
    wb, ws   = _novo_wb()
    rows_in_file = 0
    buf_final    = None

    def _salvar_parte():
        nonlocal wb, ws, file_idx, rows_in_file
        buf = io.BytesIO(); wb.save(buf)
        if zf_externo:
            fname = f"{nome}_parte{file_idx:02d}.xlsx" if split_qtd else f"{nome}.xlsx"
            zf_externo.writestr(fname, buf.getvalue())
            _log(job_id, f"   💾 [{nome}] parte {file_idx} salva ({rows_in_file:,} linhas)")
            file_idx += 1
            wb, ws = _novo_wb()
            rows_in_file = 0
        return buf

    while offset < limite:
        fetch_now = min(batch_excel, limite - offset)
        sql = (f"SELECT {cols_sql} FROM empresas "
               f"WHERE {' AND '.join(wheres)} ORDER BY data_abertura {order_dir} "
               f"LIMIT %s OFFSET %s")
        for attempt in range(1, 4):
            try:
                c = get_conn(); cur = c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur.execute(sql, params + [fetch_now, offset])
                batch = [dict(r) for r in cur.fetchall()]
                cur.close(); c.close(); break
            except Exception as e:
                _log(job_id, f"   ⚠ DB tentativa {attempt}/3: {e}")
                time.sleep(5)
        else:
            offset += fetch_now; continue

        if not batch: break
        st["lidos"] += len(batch)

        if usar_limpeza:
            # 1. Raiz
            if usar_raiz and raiz_set:
                antes = len(batch)
                batch = [r for r in batch if re.sub(r'\D','',str(r.get('cnpj',''))).zfill(14) not in raiz_set]
                st["removidos_raiz"] += antes - len(batch)

            if not batch:
                offset += fetch_now; continue

            # 2. Enriquecimento em sub-lotes de 30k
            extras = {}
            if usar_enrich and batch:
                cnpjs = [re.sub(r'\D','',str(r.get('cnpj',''))).zfill(14) for r in batch]
                ENRICH_BATCH = 30_000
                for i in range(0, len(cnpjs), ENRICH_BATCH):
                    sub = cnpjs[i:i+ENRICH_BATCH]
                    for attempt in range(1, 4):
                        try:
                            ce = get_ext_conn()
                            cur_e = ce.cursor()
                            cur_e.execute("""
                                SELECT e.cnpj, array_agg(t.numero ORDER BY t.id)
                                FROM empresas e
                                JOIN telefones t ON e.id = t.empresa_id
                                WHERE e.cnpj = ANY(%s) GROUP BY e.cnpj
                            """, (sub,))
                            for r in cur_e.fetchall():
                                extras[r[0]] = r[1]
                            cur_e.close(); ce.close()
                            _log(job_id, f"   [{nome}] enriquecimento {min(i+ENRICH_BATCH, len(cnpjs)):,}/{len(cnpjs):,} CNPJs...")
                            break
                        except Exception as ex:
                            _log(job_id, f"   ⚠ Enriquecimento tentativa {attempt}/3: {ex}")
                            time.sleep(5)

            # 3. Monta fones + remove fixos
            all_fones_flat = []
            for row in batch:
                fones = []
                for campo in ['telefone_principal','telefone_secundario']:
                    v = re.sub(r'\D','',str(row.get(campo,'') or ''))
                    if v:
                        if usar_fixos and _eh_fixo(v):
                            st["fixos_removidos"] += 1; continue
                        fones.append(inserir9(v))
                if usar_enrich:
                    cnpj_norm = re.sub(r'\D','',str(row.get('cnpj',''))).zfill(14)
                    antes_fones = len(fones)
                    for num in extras.get(cnpj_norm, []):
                        v = re.sub(r'\D','',str(num or ''))
                        if v:
                            if usar_fixos and _eh_fixo(v):
                                st["fixos_removidos"] += 1; continue
                            f = inserir9(v)
                            if f not in fones: fones.append(f)
                    novos = len(fones) - antes_fones
                    if novos > 0:
                        st["enriquecidos"] += 1
                        st["fones_extras"]  += novos
                row['_fones_limpos'] = fones
                all_fones_flat.extend(fones)

            # 4. Blocklist com retry
            if usar_blocklist and all_fones_flat:
                bloqueados = set()
                for i in range(0, len(all_fones_flat), 30_000):
                    lote = all_fones_flat[i:i+30_000]
                    for attempt in range(1, 4):
                        try:
                            cb = get_ext_conn()
                            cur_b = cb.cursor()
                            cur_b.execute("SELECT telefone FROM blocklist WHERE telefone = ANY(%s)", (lote,))
                            bloqueados.update(r[0] for r in cur_b.fetchall())
                            cur_b.close(); cb.close()
                            break
                        except Exception as ex:
                            _log(job_id, f"   ⚠ Blocklist tentativa {attempt}/3: {ex}")
                            time.sleep(5)
                for row in batch:
                    antes = len(row.get('_fones_limpos',[]))
                    row['_fones_limpos'] = [f for f in row.get('_fones_limpos',[]) if f not in bloqueados]
                    st["bloqueados"] += antes - len(row['_fones_limpos'])

        # Escreve no Excel
        for row in batch:
            if split_qtd and rows_in_file >= chunk:
                buf_final = _salvar_parte()
            dt = row.get('data_abertura')
            ano_str  = str(dt.year) if dt else ''
            data_str = dt.strftime('%d/%m/%Y') if dt else ''

            if usar_limpeza and '_fones_limpos' in row:
                fones = row['_fones_limpos']
                row['telefone_principal']  = fones[0] if len(fones) > 0 else ''
                row['telefone_secundario'] = fones[1] if len(fones) > 1 else ''

            ws.append([resolver(col, row, ano_str, data_str) for col in formato['colunas']])
            total += 1; rows_in_file += 1; st["escritos"] += 1

        offset += len(batch)
        _log(job_id, f"   [{nome}] {total:,} / {limite:,} escritos...")

    buf_final = _salvar_parte()

    # Log de stats de limpeza
    if usar_limpeza:
        _log(job_id, f"   📊 [{nome}] Limpeza:")
        if usar_raiz:     _log(job_id, f"      🗑 Raiz: {st['removidos_raiz']:,} CNPJs removidos")
        if usar_enrich:   _log(job_id, f"      📞 Enriquecimento: {st['enriquecidos']:,} linhas enriquecidas (+{st['fones_extras']:,} fones)")
        if usar_fixos:    _log(job_id, f"      📵 Fixos removidos: {st['fixos_removidos']:,} números")
        if usar_blocklist:_log(job_id, f"      🚫 Blocklist: {st['bloqueados']:,} números removidos")
        _log(job_id, f"      ✅ Total na planilha: {st['escritos']:,} linhas")
    else:
        _log(job_id, f"   ✅ [{nome}] {total:,} registros — ordem: {ordem}{f' ({file_idx-1} arquivos)' if split_qtd else ''}")

    return (None if zf_externo else buf_final.getvalue()), total


def _executar_exportacao_dual(job_id, frm, formatos):
    """Exporta dois formatos/filtros em um único ZIP."""
    _log(job_id, "🔧 Modo duplo formato — montando filtros comuns...")

    wheres = ["situacao_cadastral_cod = '02'"]
    params = []
    if frm.get('sem_mei') == 'sim':
        wheres.append("opcao_mei = 'N'")
    naturezas = frm.getlist('naturezas')
    if naturezas:
        wheres.append("natureza_juridica_cod = ANY(%s)"); params.append(naturezas)
    ufs = frm.getlist('uf')
    if ufs:
        wheres.append("estado = ANY(%s)"); params.append(ufs)
    ano_de  = frm.get('ano_de','').strip()
    ano_ate = frm.get('ano_ate','').strip()
    if ano_de and ano_de.isdigit() and ano_ate and ano_ate.isdigit():
        wheres.append("EXTRACT(YEAR FROM data_abertura) BETWEEN %s AND %s"); params += [int(ano_de), int(ano_ate)]
    elif ano_de and ano_de.isdigit():
        wheres.append("EXTRACT(YEAR FROM data_abertura) >= %s"); params.append(int(ano_de))
    elif ano_ate and ano_ate.isdigit():
        wheres.append("EXTRACT(YEAR FROM data_abertura) <= %s"); params.append(int(ano_ate))
    cnaes_raw = frm.get('cnaes','').strip()
    if cnaes_raw:
        lista = [c.strip() for c in cnaes_raw.split(',') if c.strip()]
        if lista: wheres.append("atividade_principal_cod = ANY(%s)"); params.append(lista)
    if frm.get('com_telefone') == 'sim':
        wheres.append("telefone_principal IS NOT NULL AND telefone_principal != ''")
    if frm.get('com_email') == 'sim':
        wheres.append("email IS NOT NULL AND email != ''")
    if frm.get('so_disponiveis') == 'sim':
        wheres.append("EXISTS (SELECT 1 FROM limpeza_api l WHERE l.cnpj = empresas.cnpj AND l.status = 'disponivel')")

    batch_excel = int(frm.get('batch_excel','') or 200_000)
    zip_nome    = re.sub(r'[^\w\-]', '_', frm.get('nome_lista','').strip()) or "listas_duplas"
    modo_pct    = frm.get('modo_divisao','qtd') == 'pct'
    split_on    = frm.get('split_ativo','nao') == 'sim'
    split_qtd   = int(frm.get('split_qtd','') or 0) if split_on else None

    # Limpeza
    usar_limpeza   = frm.get('limpeza_ativo','nao') == 'sim'
    usar_raiz      = frm.get('limpeza_raiz','nao') == 'sim'
    usar_enrich    = frm.get('limpeza_enrich','nao') == 'sim'
    usar_fixos     = frm.get('limpeza_fixos','nao') == 'sim'
    usar_blocklist = frm.get('limpeza_blocklist','nao') == 'sim'

    raiz_set = set()
    if usar_limpeza and usar_raiz:
        _log(job_id, "📋 Carregando raiz...")
        try:
            cr = get_ext_conn()
            cur_r = cr.cursor()
            cur_r.execute("SELECT cnpj FROM raiz_cnpjs")
            raiz_set = {r[0] for r in cur_r.fetchall()}
            cur_r.close(); cr.close()
            _log(job_id, f"   ✓ {len(raiz_set):,} CNPJs na raiz")
        except Exception as e:
            _log(job_id, f"   ⚠ Raiz: {e}")

    if usar_limpeza:
        etapas = []
        if usar_raiz: etapas.append('raiz')
        if usar_enrich: etapas.append('enriquecimento')
        if usar_fixos: etapas.append('fixos')
        if usar_blocklist: etapas.append('blocklist')
        _log(job_id, f"🧹 Pipeline: {' → '.join(etapas) if etapas else 'nenhuma etapa'}")

    # COUNT total para calcular % se necessário
    if modo_pct:
        _log(job_id, "⏳ Contando total para calcular porcentagens...")
        c = get_conn(); cur = c.cursor()
        cur.execute(f"SELECT COUNT(*) FROM empresas WHERE {' AND '.join(wheres)}", params)
        total_base = cur.fetchone()[0]
        cur.close(); c.close()
        _log(job_id, f"📊 {total_base:,} registros na base")
    else:
        total_base = None

    zip_buf = io.BytesIO()
    zf = zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED)
    total_geral = 0

    for slot in ['a', 'b']:
        fmt_key  = frm.get(f'formato_dual_{slot}', list(formatos.keys())[0])
        fmt      = formatos.get(fmt_key, list(formatos.values())[0])
        ordem    = frm.get(f'ordem_dual_{slot}', 'recentes')
        nome_arq = re.sub(r'[^\w\-]', '_', frm.get(f'nome_dual_{slot}','').strip()) or f"lista_{slot}"

        if modo_pct:
            pct = float(frm.get(f'pct_dual_{slot}','50') or 50)
            qtd = max(1, int(total_base * pct / 100))
            _log(job_id, f"   {pct}% de {total_base:,} = {qtd:,} registros")
        else:
            qtd = int(frm.get(f'qtd_dual_{slot}','') or 500_000)

        _log(job_id, f"")
        _log(job_id, f"📗 Lista {slot.upper()}: {nome_arq} | {qtd:,} registros | {ordem} | {fmt['nome']}")

        cols_necessarias = colunas_necessarias(fmt)
        if 'data_abertura' not in cols_necessarias:
            cols_necessarias = sorted(set(cols_necessarias) | {'data_abertura'})
        cols_sql = ", ".join(cols_necessarias)

        xls_bytes, total = _gerar_xlsx_parcial(
            cols_sql, wheres, params, qtd, ordem, fmt, batch_excel, nome_arq, job_id,
            split_qtd=split_qtd, zf_externo=zf,
            usar_limpeza=usar_limpeza, usar_raiz=usar_raiz, usar_enrich=usar_enrich,
            usar_fixos=usar_fixos, usar_blocklist=usar_blocklist, raiz_set=raiz_set
        )
        if xls_bytes:  # sem split — adiciona arquivo único
            zf.writestr(f"{nome_arq}.xlsx", xls_bytes)
        total_geral += total

    zf.close()
    zip_buf.seek(0)
    _log(job_id, "")
    _log(job_id, f"🎉 Pronto! {total_geral:,} registros → {zip_nome}.zip")
    _export_status[job_id]["file"] = zip_buf.read()
    _export_status[job_id]["name"] = f"{zip_nome}.zip"
    _export_status[job_id]["mime"] = "application/zip"
    _export_status[job_id]["done"] = True


def _executar_exportacao(job_id, frm):
    try:
        formatos = carregar_formatos()

        # ── Modo duplo formato ────────────────────────────────
        if frm.get('dual_fmt_ativo') == 'sim':
            _executar_exportacao_dual(job_id, frm, formatos)
            return

        ano_de    = frm.get('ano_de','').strip()
        ano_ate   = frm.get('ano_ate','').strip()
        cnaes_raw = frm.get('cnaes','').strip()
        limite    = int(frm.get('limite','') or 50_000_000)
        com_tel        = frm.get('com_telefone','nao')
        com_email      = frm.get('com_email','nao')
        sem_mei        = frm.get('sem_mei','nao')
        so_disponiveis = frm.get('so_disponiveis','nao')
        naturezas = frm.getlist('naturezas')
        ufs       = frm.getlist('uf')
        formato_k = frm.get('formato','bernardo')
        formato   = formatos.get(formato_k, list(formatos.values())[0])
        split_on  = frm.get('split_ativo','nao') == 'sim'
        split_qtd = int(frm.get('split_qtd','') or LIMITE_AUTO)
        batch_excel = int(frm.get('batch_excel','') or 200_000)
        nome_lista = re.sub(r'[^\w\-]', '_', frm.get('nome_lista','').strip()) or f"leads_{formato_k}"

        _log(job_id, "🔧 Montando filtros...")

        wheres = ["situacao_cadastral_cod = '02'"]
        params = []

        if naturezas:
            wheres.append("natureza_juridica_cod = ANY(%s)")
            params.append(naturezas)
            _log(job_id, f"   ✓ Natureza jurídica: {len(naturezas)} selecionada(s)")
        if sem_mei == 'sim':
            wheres.append("opcao_mei = 'N'")
            _log(job_id, "   ✓ Excluindo MEI (apenas opcao_mei = N)")
        if ano_de and ano_de.isdigit() and ano_ate and ano_ate.isdigit():
            wheres.append("EXTRACT(YEAR FROM data_abertura) BETWEEN %s AND %s")
            params.append(int(ano_de))
            params.append(int(ano_ate))
            _log(job_id, f"   ✓ Período: {ano_de} até {ano_ate}")
        elif ano_de and ano_de.isdigit():
            wheres.append("EXTRACT(YEAR FROM data_abertura) >= %s")
            params.append(int(ano_de))
            _log(job_id, f"   ✓ Período: a partir de {ano_de}")
        elif ano_ate and ano_ate.isdigit():
            wheres.append("EXTRACT(YEAR FROM data_abertura) <= %s")
            params.append(int(ano_ate))
            _log(job_id, f"   ✓ Período: até {ano_ate}")
        if ufs:
            wheres.append("estado = ANY(%s)")
            params.append(ufs)
            _log(job_id, f"   ✓ Estados: {', '.join(ufs)}")
        if cnaes_raw:
            lista = [c.strip() for c in cnaes_raw.split(',') if c.strip()]
            if lista:
                wheres.append("atividade_principal_cod = ANY(%s)")
                params.append(lista)
                _log(job_id, f"   ✓ CNAEs: {', '.join(lista)}")
        if com_tel == 'sim':
            wheres.append("telefone_principal IS NOT NULL AND telefone_principal != ''")
            _log(job_id, "   ✓ Somente com telefone")
        if com_email == 'sim':
            wheres.append("email IS NOT NULL AND email != ''")
            _log(job_id, "   ✓ Somente com e-mail")
        if so_disponiveis == 'sim':
            wheres.append("EXISTS (SELECT 1 FROM limpeza_api l WHERE l.cnpj = empresas.cnpj AND l.status = 'disponivel')")
            _log(job_id, "   ✓ Somente disponíveis (limpeza API)")

        params.append(limite)

        usar_limpeza    = frm.get('limpeza_ativo','nao') == 'sim'
        usar_raiz       = frm.get('limpeza_raiz','nao') == 'sim'
        usar_enrich     = frm.get('limpeza_enrich','nao') == 'sim'
        usar_fixos      = frm.get('limpeza_fixos','nao') == 'sim'
        usar_blocklist  = frm.get('limpeza_blocklist','nao') == 'sim'

        # Busca apenas as colunas que o formato realmente usa
        cols_necessarias = colunas_necessarias(formato)
        if usar_limpeza:
            for f in ['telefone_principal','telefone_secundario','cnpj']:
                if f not in cols_necessarias: cols_necessarias = sorted(set(cols_necessarias)|{f})
        cols_sql = ", ".join(cols_necessarias)

        # ── Pré-carrega raiz se necessário ───────────────────────────
        raiz_set = set()
        if usar_limpeza and usar_raiz:
            _log(job_id, "📋 Carregando lista raiz...")
            try:
                cr = get_ext_conn()
                cur_r = cr.cursor()
                cur_r.execute("SELECT cnpj FROM raiz_cnpjs")
                raiz_set = {r[0] for r in cur_r.fetchall()}
                cur_r.close(); cr.close()
                _log(job_id, f"   ✓ {len(raiz_set):,} CNPJs na raiz")
            except Exception as e:
                _log(job_id, f"   ⚠ Erro ao carregar raiz: {e}")

        # _eh_fixo definida no nível do módulo

        def _aplicar_limpeza_batch(batch_rows, formato, job_id):
            # 1. Raiz
            if usar_raiz and raiz_set:
                batch_rows = [r for r in batch_rows
                              if re.sub(r'\D','',str(r.get('cnpj',''))).zfill(14) not in raiz_set]
            if not batch_rows:
                return []

            # 2. Enriquecimento append
            if usar_enrich:
                cnpjs = [re.sub(r'\D','',str(r.get('cnpj',''))).zfill(14) for r in batch_rows]
                try:
                    ce = get_ext_conn()
                    cur_e = ce.cursor()
                    cur_e.execute("""
                        SELECT e.cnpj, array_agg(t.numero ORDER BY t.id) as telefones
                        FROM empresas e
                        JOIN telefones t ON e.id = t.empresa_id
                        WHERE e.cnpj = ANY(%s)
                        GROUP BY e.cnpj
                    """, (cnpjs,))
                    extras = {r[0]: r[1] for r in cur_e.fetchall()}
                    cur_e.close(); ce.close()
                except Exception as ex:
                    _log(job_id, f"   ⚠ Enriquecimento: {ex}")
                    extras = {}
            else:
                extras = {}

            # 3. Para cada row, monta fones finais
            all_fones_flat = []
            rows_com_fones = []
            for row in batch_rows:
                row = dict(row)
                fones = []
                for campo in ['telefone_principal','telefone_secundario']:
                    v = re.sub(r'\D','',str(row.get(campo,'') or ''))
                    if v:
                        if usar_fixos and _eh_fixo(v): continue
                        f = inserir9(v)
                        if f not in fones: fones.append(f)
                if usar_enrich:
                    cnpj_norm = re.sub(r'\D','',str(row.get('cnpj',''))).zfill(14)
                    for num in extras.get(cnpj_norm, []):
                        v = re.sub(r'\D','',str(num or ''))
                        if v:
                            if usar_fixos and _eh_fixo(v): continue
                            f = inserir9(v)
                            if f not in fones: fones.append(f)
                row['_fones_limpos'] = fones
                all_fones_flat.extend(fones)
                rows_com_fones.append(row)

            # 4. Blocklist
            if usar_blocklist and all_fones_flat:
                bloqueados = set()
                BLOCK_SIZE = 30_000
                for i in range(0, len(all_fones_flat), BLOCK_SIZE):
                    lote = all_fones_flat[i:i+BLOCK_SIZE]
                    try:
                        cb = get_ext_conn()
                        cur_b = cb.cursor()
                        cur_b.execute("SELECT telefone FROM blocklist WHERE telefone = ANY(%s)", (lote,))
                        bloqueados.update(r[0] for r in cur_b.fetchall())
                        cur_b.close(); cb.close()
                    except Exception as ex:
                        _log(job_id, f"   ⚠ Blocklist: {ex}")
                for row in rows_com_fones:
                    row['_fones_limpos'] = [f for f in row['_fones_limpos'] if f not in bloqueados]

            return rows_com_fones

        # ── Lê formato para identificar colunas de fone ──────────────
        fone_campos_idx = [(i, c) for i, c in enumerate(formato['colunas'])
                           if 'telefone' in c.get('campo','') and 'num' in c.get('campo','')]

        _log(job_id, f"   ✓ Colunas selecionadas: {cols_sql}")
        if usar_limpeza:
            etapas = []
            if usar_raiz: etapas.append('raiz')
            if usar_enrich: etapas.append('enriquecimento')
            if usar_fixos: etapas.append('remove fixos')
            if usar_blocklist: etapas.append('blocklist')
            _log(job_id, f"   🧹 Pipeline: {' → '.join(etapas) if etapas else 'nenhuma etapa selecionada'}")

        sql_final = (
            f"SELECT {cols_sql}\n"
            f"FROM empresas\n"
            f"WHERE {chr(10)+'  AND '.join(wheres)}\n"
            f"LIMIT {limite}"
        )

        _log(job_id, "")
        _log(job_id, "📋 SQL gerado:")
        _log(job_id, "─" * 50)
        sql_display = sql_final
        for p in params[:-1]:
            if isinstance(p, list):
                val = "(" + ", ".join(f"'{x}'" for x in p) + ")"
                sql_display = sql_display.replace("%s", val, 1)
            else:
                sql_display = sql_display.replace("%s", f"'{p}'", 1)
        for line in sql_display.split('\n'):
            _log(job_id, line)
        _log(job_id, "─" * 50)
        _log(job_id, "")

        _log(job_id, "⏳ Contando registros...")
        conn_count = get_conn()
        cur_count  = conn_count.cursor()
        cur_count.execute(f"SELECT COUNT(*) FROM empresas WHERE {' AND '.join(wheres)}", params[:-1])
        total_count = cur_count.fetchone()[0]
        cur_count.close()
        conn_count.close()
        _log(job_id, f"📊 Total a exportar: {total_count:,} registros")
        _log(job_id, "")

        _log(job_id, "🔌 Conectando ao banco de dados...")
        conn = get_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        _log(job_id, "⏳ Executando consulta...")
        t0 = time.time()
        cur.execute(f"""
            SELECT {cols_sql}
            FROM empresas WHERE {" AND ".join(wheres)} LIMIT %s
        """, params)
        elapsed = time.time() - t0
        _log(job_id, f"✅ Query iniciada em {elapsed:.1f}s — lendo do banco em lotes de 500k, escrevendo no Excel em lotes de {batch_excel:,}")
        _log(job_id, "")

        FETCH_SIZE  = 500_000
        chunk_size  = split_qtd if split_on else LIMITE_AUTO
        auto_split  = not split_on  # split automático ao atingir 1M
        usar_zip    = split_on or total_count > LIMITE_AUTO  # ZIP se dividir ou se total > 1M
        ts          = datetime.now().strftime('%Y%m%d_%H%M%S')
        headers     = [c['header'] for c in formato['colunas']]
        hdr_font    = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        hdr_fill    = PatternFill("solid", fgColor="5b4fcf")

        def novo_wb():
            _wb = Workbook(write_only=True)
            _ws = _wb.create_sheet("Leads")
            from openpyxl.cell import WriteOnlyCell
            hdr_row = []
            for h in headers:
                c = WriteOnlyCell(_ws, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                hdr_row.append(c)
            _ws.append(hdr_row)
            return _wb, _ws

        wb, ws      = novo_wb()
        file_idx    = 1
        row_in_file = 0
        total_rows  = 0
        pending     = []   # buffer de lotes antes de escrever no Excel
        zip_buf     = io.BytesIO() if usar_zip else None
        zf          = zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) if usar_zip else None

        def flush_pending():
            nonlocal pending
            for dr in pending:
                ws.append(dr)
            pending = []

        def salvar_arquivo():
            nonlocal wb, ws, file_idx, row_in_file
            flush_pending()
            buf_tmp = io.BytesIO()
            wb.save(buf_tmp)
            if usar_zip:
                zf.writestr(f"{nome_lista}_parte{file_idx:02d}.xlsx", buf_tmp.getvalue())
                _log(job_id, f"   💾 Arquivo {file_idx} salvo ({row_in_file:,} linhas)")
                file_idx += 1
                row_in_file = 0
                wb, ws = novo_wb()
            return buf_tmp

        while True:
            batch = cur.fetchmany(FETCH_SIZE)
            if not batch:
                break
            t_batch = time.time()
            _log(job_id, f"   📥 {len(batch):,} registros do banco — processando...")

            # Aplica pipeline de limpeza se ativo
            if usar_limpeza:
                batch = _aplicar_limpeza_batch(list(batch), formato, job_id)
                if not batch:
                    _log(job_id, f"   ⚠ Lote removido inteiro após limpeza — continuando...")
                    continue

            removidos_raiz = removidos_bl = 0
            for row in batch:
                if row_in_file >= chunk_size:
                    salvar_arquivo()

                da       = row.get('data_abertura')
                ano_str  = str(da.year) if da else ""
                data_str = da.strftime('%d/%m/%Y') if da else ""

                # Se limpeza ativa, substitui fones limpos na row antes de resolver
                if usar_limpeza and '_fones_limpos' in row:
                    fones_limpos = row['_fones_limpos']
                    row = dict(row)
                    row['telefone_principal']  = fones_limpos[0] if len(fones_limpos) > 0 else ''
                    row['telefone_secundario'] = fones_limpos[1] if len(fones_limpos) > 1 else ''

                pending.append([resolver(c, row, ano_str, data_str) for c in formato['colunas']])
                row_in_file += 1
                total_rows  += 1

                if len(pending) >= batch_excel:
                    flush_pending()
                    _log(job_id, f"   ✏ {total_rows:,} linhas escritas no Excel...")

            _log(job_id, f"   ✓ Lote processado em {time.time()-t_batch:.1f}s — total: {total_rows:,}")

        cur.close()
        conn.close()

        if total_rows == 0:
            _log(job_id, "⚠ Nenhum registro encontrado com esses filtros.")
            _export_status[job_id]["done"] = True
            return

        _log(job_id, "📝 Salvando arquivo final...")
        buf_final = salvar_arquivo()

        if usar_zip:
            zf.close()
            zip_buf.seek(0)
            n_arq = file_idx - 1
            _export_status[job_id]["file"] = zip_buf.read()
            _export_status[job_id]["name"] = f"{nome_lista}.zip"
            _export_status[job_id]["mime"] = "application/zip"
        else:
            _export_status[job_id]["file"] = buf_final.getvalue()
            _export_status[job_id]["name"] = f"{nome_lista}.xlsx"
            _export_status[job_id]["mime"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        _log(job_id, f"")
        _log(job_id, f"🎉 Pronto! {total_rows:,} registros exportados.")
        _export_status[job_id]["done"] = True

    except Exception as e:
        _log(job_id, f"❌ Erro: {e}")
        _export_status[job_id]["error"] = str(e)
        _export_status[job_id]["done"] = True

@app.route('/status/<job_id>')
def status(job_id):
    from flask import jsonify
    job = _export_status.get(job_id, {})
    return jsonify({
        "log":   job.get("log", []),
        "done":  job.get("done", False),
        "error": job.get("error", ""),
        "has_file": job.get("file") is not None,
        "name":  job.get("name", ""),
    })

@app.route('/download/<job_id>')
def download(job_id):
    job = _export_status.get(job_id, {})
    if not job.get("file"):
        return "Arquivo não encontrado", 404
    buf = io.BytesIO(job["file"])
    return send_file(buf, mimetype=job["mime"],
                     as_attachment=True, download_name=job["name"])


@app.route('/salvar-formato', methods=['POST'])
def salvar_formato():
    formatos = carregar_formatos()
    frm = request.form
    key  = frm.get('formato_id','').strip().lower()
    nome = frm.get('formato_nome','').strip()
    headers = frm.getlist('col_header[]')
    campos  = frm.getlist('col_campo[]')
    manuais = frm.getlist('col_manual[]')
    if key and nome and headers:
        colunas = []
        for h, c, m in zip(headers, campos, manuais):
            col = {"header": h, "campo": c}
            if c == 'manual':
                col["valor_manual"] = m
            colunas.append(col)
        formatos[key] = {"nome": nome, "colunas": colunas}
        salvar_formatos(formatos)
    return redirect('/')

@app.route('/editar-formato/<key>')
def editar_formato(key):
    formatos = carregar_formatos()
    fmt = formatos.get(key)
    if not fmt: return redirect('/')
    return render_template_string(HTML, naturezas=NATUREZAS, ufs=UFS,
        f=filtros_default(), formatos=formatos,
        campos_json=json.dumps(CAMPOS_DISPONIVEIS),
        resultado=None,
        editar_key=key,
        editar_nome=fmt['nome'],
        editar_colunas=json.dumps(fmt['colunas']))

@app.route('/excluir-formato/<key>')
def excluir_formato(key):
    formatos = carregar_formatos()
    if key in formatos and key not in ['bernardo','olos','empresaaqui']:
        del formatos[key]
        salvar_formatos(formatos)
    return redirect('/')


# ─── API C6 ───────────────────────────────────────────────────

_api_jobs = {}  # job_id → {log, done, error}

TOKEN_URL   = "https://crm-leads-p.c6bank.info/querie-partner/token"
CONSULTA_URL= "https://crm-leads-p.c6bank.info/querie-partner/client/avaliable"

C6_CREDS = {
    "chave1": {"CLIENT_ID": "EA8ZUFeZVSeqMGr49XJSsZKFuxSZub3i", "CLIENT_SECRET": "EUomxjGf6BvBZ1HO",  "name": "Chave 1"},
    "chave2": {"CLIENT_ID": "imWzrW41HcnoJgvZqHCaLvziUGlhAJAH", "CLIENT_SECRET": "A0lAqZO73uW3wryU", "name": "Chave 2"},
}

def _api_log(job_id, msg):
    _api_jobs[job_id]["log"].append(msg)

def _c6_token(creds):
    """Obtém access_token OAuth2 da API C6."""
    import urllib.parse, urllib.request, urllib.error
    params = urllib.parse.urlencode({
        "grant_type":    "client_credentials",
        "client_id":     creds["CLIENT_ID"],
        "client_secret": creds["CLIENT_SECRET"],
    })
    req = urllib.request.Request(TOKEN_URL, data=params.encode(), method="POST")
    req.add_header("Content-Type", "application/x-www-form-urlencoded")
    req.add_header("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    req.add_header("Accept", "application/json, text/plain, */*")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read())["access_token"]
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors='replace')
        raise Exception(f"TOKEN falhou {e.code}: {body[:200]}")

def _c6_consultar(cnpjs, creds):
    """Envia lista de CNPJs para API e retorna set dos disponíveis."""
    import urllib.request, urllib.error

    # Normaliza CNPJs igual ao código JS: só dígitos, 14 chars
    cnpjs_norm = [re.sub(r'\D','',str(c)).zfill(14) for c in cnpjs]

    token = _c6_token(creds)
    body  = json.dumps({"CNPJ": cnpjs_norm}).encode()
    req   = urllib.request.Request(CONSULTA_URL, data=body, method="POST")
    req.add_header("Authorization", f"Bearer {token}")
    req.add_header("Content-Type",  "application/json")
    req.add_header("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    req.add_header("Accept", "application/json, text/plain, */*")
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.loads(r.read())
    except urllib.error.HTTPError as e:
        body_err = e.read().decode(errors='replace')
        raise Exception(f"CONSULTA falhou {e.code}: {body_err[:200]}")

    key = next((k for k in data if "cnpj" in k.lower() and isinstance(data[k], list)), None)
    if not key:
        return set()
    return {re.sub(r'\D','',str(c)).zfill(14) for c in data[key]}

def _salvar_limpeza(resultados: list):
    """Upsert em lote na tabela limpeza_api."""
    if not resultados:
        return
    raw_url = DATABASE_URL.replace("&channel_binding=require","")
    conn = psycopg2.connect(raw_url)
    cur  = conn.cursor()
    psycopg2.extras.execute_values(cur, """
        INSERT INTO limpeza_api (cnpj, status, processado_em)
        VALUES %s
        ON CONFLICT (cnpj) DO UPDATE
            SET status = EXCLUDED.status,
                processado_em = EXCLUDED.processado_em
    """, [(r["cnpj"], r["status"], datetime.now()) for r in resultados])
    conn.commit()
    cur.close()
    conn.close()

N8N_FISH_URL    = "https://n8n.upscales.com.br/webhook/2ccead38-deb8-48d0-9f44-0edccafcc026"
TELEFONES_DB_URL = "postgresql://neondb_owner:npg_ki2aKHqlnFY9@ep-quiet-night-ac2uu9kc-pooler.sa-east-1.aws.neon.tech/neondb?sslmode=require"

def _buscar_telefones_externos(cnpj: str) -> list:
    """Busca telefones na tabela telefones do banco externo pelo CNPJ."""
    try:
        conn = get_ext_conn()
        cur  = conn.cursor()
        cur.execute("""
            SELECT t.numero FROM telefones t
            JOIN empresas e ON e.id = t.empresa_id
            WHERE e.cnpj = %s
            LIMIT 5
        """, (cnpj,))
        rows = [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
        return rows
    except Exception:
        return []

def _fish_enviar(cnpj, row, connector, job_id):
    """
    Envia cliente ao N8N.
    Regras:
    - Aplica inserir9 em todos os telefones
    - Se não tiver telefone na empresa, busca no banco externo (telefones quentes)
    - Se não tiver nenhum telefone em nenhum lugar, NÃO envia
    """
    try:
        import urllib.parse, urllib.request

        cnpj_norm = re.sub(r"\D","",str(cnpj)).zfill(14)

        # Coleta fones da empresa aplicando inserir9
        fones = []
        for campo in ["telefone_principal", "telefone_secundario"]:
            v = re.sub(r"\D","",str(row.get(campo,"") or ""))
            if v:
                fones.append(inserir9(v))

        # Sem fone na empresa → busca no banco externo
        if not fones:
            externos = _buscar_telefones_externos(cnpj_norm)
            for num in externos:
                v = re.sub(r"\D","",str(num or ""))
                if v:
                    fones.append(inserir9(v))

        # Sem nenhum telefone em lugar nenhum → não envia
        if not fones:
            return False

        params = {
            "cpf":    cnpj_norm,
            "nome":   row.get("razao_social","") or "",
            "chave":  row.get("email","") or "",
        }
        if connector: params["conector"] = connector

        for i, fone in enumerate(fones[:4], 1):
            params[f"fone{i}"] = fone

        qs = urllib.parse.urlencode({k:v for k,v in params.items() if v})
        urllib.request.urlopen(urllib.request.Request(f"{N8N_FISH_URL}?{qs}"), timeout=10)
        return True
    except Exception as e:
        _api_log(job_id, f"  🐟 FISH ERRO: {e}")
        return False


def _executar_api(job_id, params):
    try:
        preset          = params.get("preset","ativos")
        ano_de          = str(params.get("ano_de","")).strip()
        ano_ate         = str(params.get("ano_ate","")).strip()
        cnaes_raw       = params.get("cnaes","").strip()
        limite          = int(params.get("limite","") or 0)  # 0 = sem limite
        naturezas       = params.get("naturezas",[])
        key_mode        = params.get("key_mode","chave1")
        # Batch fixo: dupla=40k (20k por chave), qualquer outra=20k
        batch_size      = 40000 if key_mode == "dupla" else 20000
        delay_min       = float(params.get("delay_min","") or 2)
        fish_mode        = params.get("fish_mode", False)
        connector        = params.get("connector","")
        fish_split       = params.get("fish_split", False)
        fish_connector_a = params.get("fish_connector_a","")
        fish_connector_b = params.get("fish_connector_b","")
        fish_pct_a       = int(params.get("fish_pct_a", 50))
        if fish_mode:
            _api_log(job_id, f"🐟 FISH params: split={fish_split} | A='{fish_connector_a}' | B='{fish_connector_b}' | pct_a={fish_pct_a}")
        reprocessar       = params.get("reprocessar", False)
        so_disponiveis_api = params.get("so_disponiveis_api", False)
        excluir_cliente = params.get("excluir_cliente", False)
        extrair_clientes= params.get("extrair_clientes", False)

        ano_atual = datetime.now().year
        if preset == "recentes":   ano_de = str(ano_atual-2)
        elif preset == "ei_recente":
            ano_de = str(ano_atual-3)
            if not naturezas: naturezas = ["2135","2305","2313"]
        elif preset == "ltda":
            if not naturezas: naturezas = ["2062","2232","2240","2259","2267","2070","2089"]

        # MEI SEMPRE EXCLUÍDO
        wheres  = ["situacao_cadastral_cod = '02'", "opcao_mei = 'N'"]
        qparams = []

        if naturezas:
            wheres.append("natureza_juridica_cod = ANY(%s)")
            qparams.append(naturezas)
        if ano_de and ano_de.isdigit() and ano_ate and ano_ate.isdigit():
            wheres.append("EXTRACT(YEAR FROM data_abertura) BETWEEN %s AND %s")
            qparams += [int(ano_de), int(ano_ate)]
        elif ano_de and ano_de.isdigit():
            wheres.append("EXTRACT(YEAR FROM data_abertura) >= %s")
            qparams.append(int(ano_de))
        elif ano_ate and ano_ate.isdigit():
            wheres.append("EXTRACT(YEAR FROM data_abertura) <= %s")
            qparams.append(int(ano_ate))
        if cnaes_raw:
            lista = [c.strip() for c in cnaes_raw.split(",") if c.strip()]
            if lista:
                wheres.append("atividade_principal_cod = ANY(%s)")
                qparams.append(lista)

        if so_disponiveis_api:
            wheres.append("EXISTS (SELECT 1 FROM limpeza_api l WHERE l.cnpj = empresas.cnpj AND l.status = 'disponivel')")
            reprocessar = True  # já estão na tabela, precisa reprocessar
            _api_log(job_id, "✅ SOMENTE DISPONÍVEIS — reconfirmando CNPJs classificados como disponível")
        elif not reprocessar:
            wheres.append("cnpj NOT IN (SELECT cnpj FROM limpeza_api)")
        else:
            _api_log(job_id, "🔄 REPROCESSAR ativo")

        select_cols = "cnpj" if not fish_mode else "cnpj,razao_social,telefone_principal,telefone_secundario,email"
        where_sql   = ' AND '.join(wheres)


        _api_log(job_id, "⏳ Contando CNPJs pendentes...")
        conn = get_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        if not reprocessar and not so_disponiveis_api:
            where_sql_count = where_sql.replace(
                "cnpj NOT IN (SELECT cnpj FROM limpeza_api)",
                "NOT EXISTS (SELECT 1 FROM limpeza_api l WHERE l.cnpj = empresas.cnpj)"
            )
        else:
            where_sql_count = where_sql

        cur.execute(f"SELECT COUNT(*) as n FROM empresas WHERE {where_sql_count}", qparams)
        total_count = cur.fetchone()["n"]
        cur.close()
        _api_log(job_id, f"📊 {total_count:,} CNPJs para processar")

        if total_count == 0:
            conn.close()
            _api_log(job_id, "⚠ Nenhum CNPJ novo.")
            _api_jobs[job_id].update({"done":True,"stats":{"total":0,"disp":0,"cli":0,"fish":0,"lote_atual":0,"lotes_total":0,"restam_cnpjs":0}})
            return

        # Store live settings
        _api_jobs[job_id]["live"] = {"key_mode":key_mode,"batch_size":batch_size,"delay_min":delay_min,"paused":False}

        lotes_total_est = (total_count + batch_size - 1) // batch_size
        _api_log(job_id, f"📦 ~{lotes_total_est:,} lotes de {batch_size:,} CNPJs")
        if fish_mode:
            if fish_split and fish_connector_a and fish_connector_b:
                _api_log(job_id, f"🐟 FISH ativo — {fish_pct_a}% → {fish_connector_a} | {100-fish_pct_a}% → {fish_connector_b}")
            else:
                _api_log(job_id, f"🐟 FISH ativo — conector: {connector or 'sem conector'}")
        _api_log(job_id, "📡 Iniciando processamento em lotes...")

        total_disp = total_cli = total_fish = 0
        MAX_RETRY  = 5
        lote_num   = 0
        offset     = 0  # tracks position — new connection each batch, no long-lived cursor
        stream_sql = f"SELECT {select_cols} FROM empresas WHERE {where_sql_count}"
        conn.close()    # close COUNT connection, we'll open fresh per batch

        while True:
            live = _api_jobs[job_id]["live"]

            if live.get("paused"):
                _api_log(job_id, "⏸ PAUSADO — aguardando retomada...")
                while _api_jobs[job_id]["live"].get("paused"):
                    time.sleep(1)
                    if _api_jobs[job_id].get("cancelar"): break
                _api_log(job_id, "▶ RETOMADO")

            if _api_jobs[job_id].get("cancelar"):
                _api_log(job_id, "🛑 Processamento cancelado.")
                break

            cur_batch = int(live.get("batch_size", batch_size))
            cur_delay = float(live.get("delay_min", delay_min))
            cur_key   = live.get("key_mode", key_mode)

            # Fresh connection per batch — survives SSL drops and idle timeouts
            fetch_ok = False
            rows = []
            for db_attempt in range(1, 4):
                try:
                    c = get_conn()
                    cur = c.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                    cur.execute(f"{stream_sql} LIMIT %s OFFSET %s", qparams + [cur_batch, offset])
                    rows = cur.fetchall()
                    cur.close(); c.close()
                    fetch_ok = True
                    break
                except Exception as db_err:
                    _api_log(job_id, f"  ⚠ DB erro (tentativa {db_attempt}/3): {db_err} — reconectando...")
                    time.sleep(5)

            if not fetch_ok:
                _api_log(job_id, "❌ Falha ao buscar lote do banco após 3 tentativas, pulando.")
                offset += cur_batch
                continue

            if not rows:
                break  # acabou

            offset   += len(rows)
            lote_num += 1
            restam    = max(0, total_count - offset)
            lotes_total_est = lote_num + (restam + cur_batch - 1) // cur_batch if restam else lote_num

            cnpj_list = [r["cnpj"] if isinstance(r,dict) else r[0] for r in rows]
            row_map   = {(r["cnpj"] if isinstance(r,dict) else r[0]): r for r in rows} if fish_mode else {}

            _api_log(job_id, f"=== Lote {lote_num}/{lotes_total_est} ({len(cnpj_list):,} CNPJs | {cur_key} | delay:{cur_delay}min) ===")
            disponiveis = None

            if cur_key == "dupla":
                meio = len(cnpj_list)//2
                for attempt in range(1, MAX_RETRY+1):
                    try:
                        d1 = _c6_consultar(cnpj_list[:meio], C6_CREDS["chave1"])
                        d2 = _c6_consultar(cnpj_list[meio:], C6_CREDS["chave2"])
                        disponiveis = d1 | d2; break
                    except Exception as e:
                        _api_log(job_id, f"  ⚠ Tentativa {attempt}/{MAX_RETRY}: {e}")
                        if attempt < MAX_RETRY: time.sleep(120)
            else:
                creds = C6_CREDS["chave2"] if cur_key=="chave2" else                         (C6_CREDS["chave1"] if lote_num%2==1 else C6_CREDS["chave2"]) if cur_key=="intercalar" else                         C6_CREDS["chave1"]
                for attempt in range(1, MAX_RETRY+1):
                    try:
                        disponiveis = _c6_consultar(cnpj_list, creds); break
                    except Exception as e:
                        _api_log(job_id, f"  ⚠ Tentativa {attempt}/{MAX_RETRY}: {e}")
                        if attempt < MAX_RETRY: time.sleep(120)

            if disponiveis is None:
                _api_log(job_id, f"  ❌ Lote {lote_num} falhou, pulando.")
                continue

            resultados = []
            for cnpj in cnpj_list:
                cnpj_norm = re.sub(r"\D","",str(cnpj)).zfill(14)
                status = "disponivel" if cnpj_norm in disponiveis else "cliente"
                resultados.append({"cnpj":cnpj_norm,"status":status})

            disp = sum(1 for r in resultados if r["status"]=="disponivel")
            cli  = len(resultados) - disp
            total_disp += disp; total_cli += cli
            _salvar_limpeza(resultados)
            _api_log(job_id, f"  ✅ {disp:,} disponível | {cli:,} cliente | restam ~{restam:,}")

            if fish_mode:
                fish_log  = _api_jobs[job_id].setdefault("fish_log", [])
                fish_dist = _api_jobs[job_id].setdefault("fish_dist", {})
                for r in resultados:
                    if r["status"] == "cliente":
                        if fish_split and fish_connector_a and fish_connector_b:
                            # Distribui corretamente: compara proporção atual com alvo
                            cnt_a = fish_dist.get(fish_connector_a, 0)
                            cnt_b = fish_dist.get(fish_connector_b, 0)
                            total_ab = cnt_a + cnt_b
                            atual_a  = cnt_a / max(total_ab, 1)
                            target_a = fish_pct_a / 100
                            conn_usar = fish_connector_a if atual_a <= target_a else fish_connector_b
                        else:
                            conn_usar = connector
                        row_data = row_map.get(r["cnpj"],{})
                        enviado = _fish_enviar(r["cnpj"], row_data, conn_usar, job_id)
                        if enviado:
                            total_fish += 1
                            fish_dist[conn_usar or 'sem conector'] = fish_dist.get(conn_usar or 'sem conector', 0) + 1
                            fones = []
                            for campo in ["telefone_principal","telefone_secundario"]:
                                v = re.sub(r"\D","",str(row_data.get(campo,"") or ""))
                                if v: fones.append(inserir9(v))
                            fish_log.append({
                                "cnpj": r["cnpj"],
                                "nome": (row_data.get("razao_social","") or "")[:40],
                                "fones": fones[:4],
                                "conector": conn_usar or 'sem conector',
                            })
                if total_fish: _api_log(job_id, f"  🐟 Fish: {total_fish:,} enviados | {dict(fish_dist)}")

            _api_jobs[job_id]["stats"] = {"total":total_disp+total_cli,"disp":total_disp,"cli":total_cli,"fish":total_fish,
                                           "lote_atual":lote_num,"lotes_total":lotes_total_est,"restam_cnpjs":restam}

            if restam > 0:
                elapsed = 0
                while elapsed < cur_delay * 60:
                    time.sleep(1); elapsed += 1
                    if _api_jobs[job_id].get("cancelar"): break
                    if _api_jobs[job_id]["live"].get("paused"): break

        _api_log(job_id, "")
        _api_log(job_id, f"🎉 Concluído! {total_disp:,} disponível | {total_cli:,} cliente | {total_fish:,} fish")

        # Libera chaves
        _api_log(job_id, f"🔓 Chave(s) liberadas: {', '.join(keys_locked)}")

        # Fish final batch (caso haja clientes ainda não enviados — o loop envia por lote,
        # mas se excluir_cliente estiver ativo garantimos que TODOS foram enviados antes de deletar)
        if fish_mode and excluir_cliente and total_cli > 0:
            _api_log(job_id, "🐟 Confirmando envio Fish antes de excluir...")

        if excluir_cliente and total_cli > 0:
            _api_log(job_id, "🗑 Excluindo clientes do banco...")
            try:
                conn2 = get_conn(); cur2 = conn2.cursor()
                cur2.execute("DELETE FROM empresas WHERE cnpj IN (SELECT cnpj FROM limpeza_api WHERE status='cliente')")
                deleted = cur2.rowcount; conn2.commit(); cur2.close(); conn2.close()
                _api_log(job_id, f"✅ {deleted:,} registros removidos")
            except Exception as e: _api_log(job_id, f"❌ Erro ao excluir: {e}")

        if extrair_clientes and total_cli > 0:
            _api_log(job_id, "📥 Gerando Excel de clientes...")
            try:
                conn3 = get_conn(); cur3 = conn3.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
                cur3.execute("""SELECT e.cnpj,e.razao_social,e.telefone_principal,e.telefone_secundario,
                    e.email,e.estado,e.cidade,e.atividade_principal_cod,l.processado_em
                    FROM limpeza_api l JOIN empresas e ON e.cnpj=l.cnpj
                    WHERE l.status='cliente' ORDER BY l.processado_em DESC""")
                clientes = cur3.fetchall(); cur3.close(); conn3.close()
                wb = Workbook(); ws = wb.active; ws.title = "Clientes"
                cols_xls = ["cnpj","razao_social","telefone_principal","telefone_secundario","email","estado","cidade","atividade_principal_cod","processado_em"]
                ws.append(cols_xls)
                for r in clientes: ws.append([str(r.get(col,"") or "") for col in cols_xls])
                xls_buf = io.BytesIO(); wb.save(xls_buf)
                _api_jobs[job_id]["clientes_xlsx"] = xls_buf.getvalue()
                _api_jobs[job_id]["clientes_count"] = len(clientes)
                _api_log(job_id, f"✅ Excel gerado: {len(clientes):,} clientes")
            except Exception as e: _api_log(job_id, f"❌ Erro ao extrair: {e}")

        _api_jobs[job_id].update({"done":True,"stats":{"total":total_disp+total_cli,"disp":total_disp,"cli":total_cli,"fish":total_fish}})

    except Exception as e:
        _api_log(job_id, f"❌ Erro fatal: {e}")
        _api_jobs[job_id]["error"] = str(e)

        if _api_jobs[job_id].get("cancelar"):
            _api_jobs[job_id]["done"] = True
            return

        # Auto-retry: aguarda e recomeça do ponto onde parou
        # (CNPJs já salvos em limpeza_api são automaticamente pulados)
        retry_count = _api_jobs[job_id].get("_retry_count", 0) + 1
        _api_jobs[job_id]["_retry_count"] = retry_count
        _api_log(job_id, f"🔄 Retentativa {retry_count} em 5 min (continuando de onde parou)...")
        _api_jobs[job_id]["error"] = ""

        for _ in range(300):  # 5 min fixo
            time.sleep(1)
            if _api_jobs[job_id].get("cancelar"):
                _api_log(job_id, "🛑 Cancelado durante espera de retentativa.")
                _api_jobs[job_id]["done"] = True
                return

        _api_log(job_id, f"▶ Retomando processamento...")
        _executar_api(job_id, params)  # chama recursivamente — retoma do ponto atual


@app.route('/api/iniciar', methods=['POST'])
def api_iniciar():
    from flask import jsonify
    params   = request.get_json()
    key_mode = params.get("key_mode", "chave1")

    job_id = str(uuid.uuid4())[:8]
    _api_jobs[job_id] = {"log": [], "done": False, "error": ""}
    t = threading.Thread(target=_executar_api, args=(job_id, params))
    t.daemon = True
    t.start()
    return jsonify({"job_id": job_id})

@app.route('/api/status/<job_id>')
def api_status(job_id):
    from flask import jsonify
    job = _api_jobs.get(job_id, {"log":[], "done":True, "error":"não encontrado"})
    return jsonify({"log":job.get("log",[]),"done":job.get("done",True),"error":job.get("error",""),
                    "stats":job.get("stats",{}),"has_clientes_xlsx":"clientes_xlsx" in job,
                    "clientes_count":job.get("clientes_count",0),
                    "paused":job.get("live",{}).get("paused",False),
                    "live":job.get("live",{})})

@app.route('/api/controle/<job_id>', methods=['POST'])
def api_controle(job_id):
    from flask import jsonify
    job = _api_jobs.get(job_id)
    if not job: return jsonify({"ok":False}), 404
    data   = request.get_json()
    action = data.get("action","")
    live   = job.setdefault("live", {})
    if   action == "pause":  live["paused"] = True
    elif action == "resume": live["paused"] = False
    elif action == "cancel":
        job["cancelar"] = True; live["paused"] = False
    elif action == "update":
        if "key_mode" in data:
            live["key_mode"]   = data["key_mode"]
            live["batch_size"] = 40000 if data["key_mode"] == "dupla" else 20000
        if "delay_min" in data: live["delay_min"] = float(data["delay_min"])
    return jsonify({"ok":True,"live":live})

@app.route('/api/teste-fish', methods=['POST'])
def api_teste_fish():
    from flask import jsonify
    data = request.get_json()
    cnpj = re.sub(r'\D','',str(data.get('cnpj','')).strip()).zfill(14)
    connector = data.get('connector','')
    if len(cnpj) != 14:
        return jsonify({"erro": "CNPJ inválido — precisa ter 14 dígitos"})
    # Busca empresa no banco principal
    try:
        conn = get_conn()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT e.cnpj, e.razao_social, e.telefone_principal, e.telefone_secundario,
                   e.email, e.cidade, e.estado,
                   l.status as status_api
            FROM empresas e
            LEFT JOIN limpeza_api l ON l.cnpj = e.cnpj
            WHERE e.cnpj = %s
        """, (cnpj,))
        row = cur.fetchone()
        cur.close(); conn.close()
    except Exception as e:
        return jsonify({"erro": f"Erro ao consultar banco: {e}"})
    if not row:
        return jsonify({"erro": f"CNPJ {cnpj} não encontrado na base"})
    row = dict(row)
    # Coleta telefones
    fones = []
    fones_source = 'empresa'
    for campo in ['telefone_principal', 'telefone_secundario']:
        v = re.sub(r'\D','',str(row.get(campo,'') or ''))
        if v: fones.append(inserir9(v))
    if not fones:
        externos = _buscar_telefones_externos(cnpj)
        fones = [inserir9(re.sub(r'\D','',str(n))) for n in externos if re.sub(r'\D','',str(n))]
        if fones: fones_source = 'externo'
    # Monta payload
    payload = {
        "cpf":  cnpj,
        "nome": row.get('razao_social','') or '',
        "chave": row.get('email','') or '',
    }
    if connector: payload["conector"] = connector
    for i, fone in enumerate(fones[:4], 1):
        payload[f"fone{i}"] = fone
    payload = {k:v for k,v in payload.items() if v}
    return jsonify({
        "empresa": {
            "cnpj": row['cnpj'],
            "razao_social": row.get('razao_social',''),
            "email": row.get('email',''),
            "cidade": row.get('cidade',''),
            "estado": row.get('estado',''),
            "status_api": row.get('status_api',''),
        },
        "fones": fones,
        "fones_source": fones_source,
        "payload": payload,
    })

@app.route('/api/teste-fish-enviar', methods=['POST'])
def api_teste_fish_enviar():
    from flask import jsonify
    import urllib.parse, urllib.request
    payload = request.get_json()
    try:
        qs  = urllib.parse.urlencode({k:v for k,v in payload.items() if v})
        req = urllib.request.Request(f"{N8N_FISH_URL}?{qs}", method="GET")
        req.add_header("User-Agent","Mozilla/5.0")
        urllib.request.urlopen(req, timeout=10)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)})

@app.route('/api/fish-log/<job_id>')
def api_fish_log(job_id):
    from flask import jsonify
    job = _api_jobs.get(job_id, {})
    return jsonify({"log": job.get("fish_log",[]), "dist": job.get("fish_dist",{})})

@app.route('/api/resultados')
def api_resultados():
    from flask import jsonify
    resultado = request.args.get("resultado","")
    limite    = int(request.args.get("limite","100"))
    conn = get_conn()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT COUNT(*) as n FROM empresas WHERE situacao_cadastral_cod='02' AND opcao_mei='N'")
    total_bd    = cur.fetchone()["n"]
    cur.execute("SELECT COUNT(*) as n FROM limpeza_api")
    total       = cur.fetchone()["n"]
    cur.execute("SELECT COUNT(*) as n FROM limpeza_api WHERE status='disponivel'")
    disponiveis = cur.fetchone()["n"]
    cur.execute("SELECT COUNT(*) as n FROM limpeza_api WHERE status='cliente'")
    clientes    = cur.fetchone()["n"]

    # Naturezas: total, processado, disponivel, cliente por natureza
    cur.execute("""
        SELECT
            e.natureza_juridica_cod AS cod,
            e.natureza_juridica     AS nome,
            COUNT(*)                AS total,
            COUNT(l.cnpj)           AS processados,
            SUM(CASE WHEN l.status='disponivel' THEN 1 ELSE 0 END) AS disponiveis,
            SUM(CASE WHEN l.status='cliente'    THEN 1 ELSE 0 END) AS clientes
        FROM empresas e
        LEFT JOIN limpeza_api l ON l.cnpj = e.cnpj
        WHERE e.situacao_cadastral_cod = '02' AND e.opcao_mei = 'N'
        GROUP BY e.natureza_juridica_cod, e.natureza_juridica
        ORDER BY processados DESC
        LIMIT 15
    """)
    naturezas_stats = [dict(r) for r in cur.fetchall()]

    where = f"WHERE status='{resultado}'" if resultado else ""
    cur.execute(f"SELECT cnpj,status,processado_em FROM limpeza_api {where} ORDER BY processado_em DESC LIMIT %s",(limite,))
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        if r.get("processado_em"): r["processado_em"] = r["processado_em"].strftime("%d/%m/%Y %H:%M")
    cur.close(); conn.close()
    return jsonify({"total":total,"total_bd":total_bd,"disponiveis":disponiveis,"clientes":clientes,
                    "rows":rows,"naturezas_stats":naturezas_stats})


@app.route('/api/exportar-resultados')
def api_exportar_resultados():
    resultado = request.args.get("resultado","")
    conn = get_conn()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = f"WHERE l.status='{resultado}'" if resultado else ""
    cur.execute(f"""
        SELECT l.cnpj, l.status, l.processado_em,
               e.razao_social, e.telefone_principal, e.telefone_secundario,
               e.email, e.estado, e.cidade, e.atividade_principal_cod
        FROM limpeza_api l
        LEFT JOIN empresas e ON e.cnpj = l.cnpj
        {where}
        ORDER BY l.processado_em DESC
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    headers = ["cnpj","status","processado_em","razao_social","telefone_principal",
               "telefone_secundario","email","estado","cidade","atividade_principal_cod"]
    ws.append(headers)
    for row in rows:
        ws.append([str(row.get(h,"") or "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"limpeza_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@app.route('/api/download-clientes/<job_id>')
def api_download_clientes(job_id):
    job = _api_jobs.get(job_id, {})
    if "clientes_xlsx" not in job: return "Não disponível", 404
    buf = io.BytesIO(job["clientes_xlsx"])
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

if __name__ == '__main__':
    print("Servidor iniciado em http://localhost:5000")
    app.run(debug=False, port=5000)